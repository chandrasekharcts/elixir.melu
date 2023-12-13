from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import user_details, account_list, projects_list, sub_process_list, update_type_list, update_source_list, \
    sub_process_details, update_details, update_recipients, question_list, question_details, conquest_list, \
    conquest_questions, conquest_recipients, conquest_recipients_answers, update_conquest
from django.contrib.auth.models import User
from dashboard.forms import update_editorForm, conquest_questionForm
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.message import EmailMessage
from email import encoders
from itertools import chain
import smtplib, json, os, datetime, xlrd, re, ldap
from elixir_mleu import settings
from django.template.loader import render_to_string
from django.core.files.storage import FileSystemStorage
import random
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.db.models import Q
import time
import openpyxl





# -------------------------- Automatic Account Login URL Redirection ---------------------------------------------------
def userRedirrectLogin(request):
    response = redirect('/accounts/login/')
    return response
# --------------------------------- END Function -----------------------------------------------------------------------





# --------------------------------- User Login Authentication ----------------------------------------------------------


# --------------------------------- LOGIN AUTHENTICATION ---------------------------------------------------------------
def user_login(request):
    if request.user.is_authenticated:
        request.session.set_expiry(3600)
        return redirect('/dashboard/')
    else:
        if request.method == 'POST':
            emp_id = request.POST.get('username')
            username = fetch_username(emp_id)
            password = request.POST.get('password')
            if username == "invalid":
                return render(request, 'a1-login.html',
                              {'msg': "This Employee ID user is not registered with this tool."})
            else:
                user = authenticate(username=username, password="cognizant")
                if user:
                    if user.is_active:
                        #print("ldap Checking Started")
                        #check = check_credentials(emp_id, password)
                        #print("ldap Checking Completed")
                        check = "Success"
                        if check == "Success":
                            login(request, user)
                            request.session.set_expiry(3600)
                            return redirect('/dashboard/')
                        else:
                            return render(request, 'a1-login.html',
                                          {'msg': check})
                    else:
                        return render(request, 'a1-login.html',
                                      {'msg': "Your account is inactive, Please contact your administrator"})
                else:
                    return render(request, 'a1-login.html', {'msg': "Invalid login Credentials"})
        else:
            return render(request, 'a1-login.html', {})
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LDAP AUTHENTICATION ----------------------------------------------------------------
def check_credentials(username, password):
    """Verifies credentials for username and password.
    Returns None on success or a string describing the error on failure
    # Adapt to your needs
    """
    #LDAP_SERVER = 'ldap://10.242.175.105:389'
    LDAP_SERVER = 'ldap://10.127.82.92:389'
    # fully qualified AD user name
    LDAP_USERNAME = '%s@cognizant.com' % username
    # your password
    LDAP_PASSWORD = password
    base_dn = 'DC=cts,DC=com'
    ldap_filter = 'userPrincipalName=%s@cognizant.com' % username
    attrs = ['memberOf']
    try:
        # build a client
        ldap_client = ldap.initialize(LDAP_SERVER)
        # perform a synchronous bind
        ldap_client.set_option(ldap.OPT_REFERRALS, 0)
        ldap_client.simple_bind_s(LDAP_USERNAME, LDAP_PASSWORD)
    except ldap.INVALID_CREDENTIALS:
        ldap_client.unbind()
        return 'Invalid username or password. Please try again!!'
    except ldap.SERVER_DOWN:
        return 'AD server not available. Please try again!!'
    return "Success"
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FETCH USERNAME FROM THE EMP ID -----------------------------------------------------
def fetch_username(emp_id):
    username = user_details.objects.filter(emp_id=emp_id)
    if username:
        return username[0].email_id
    else:
        return "invalid"
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Function------------------------------------------------------------------------





# --------------------------------- User  ------------------------------------------------------------------------
@login_required
def user_logout(request):
    request.session.set_expiry(0)
    logout(request)
    return render(request, 'g1-logout.html')
# --------------------------------- END Function -----------------------------------------------------------------------





# --------------------------------- VERSION HISTORY --------------------------------------------------------------------
@login_required
def version_history(request):
    request.session.set_expiry(0)
    return render(request, 'g2-version-history.html')
# --------------------------------- END Function -----------------------------------------------------------------------





# --------------------------------- DASHBOARD --------------------------------------------------------------------------


# --------------------------------- REDIRECT TO DASHBOARD --------------------------------------------------------------
@login_required
def dashbaord(request):
    request.session.set_expiry(3600)
    return render(request, 'b1-dashboard.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LOAD USER RELATED ACCOUNTS LIST ----------------------------------------------------
def user_account_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_type = request.user.email

        data['acc_name'] = user_acc_list(user_type, request.user.username)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)

def user_acc_list(user_type, user_name):
    accounts = []
    if user_type == "superuser" or user_type == "leader":
        acc_list = account_list.objects.all().values('account_name')
        for acc in acc_list:
            set = {}
            set['account_name'] = acc['account_name']
            accounts.append(set)

    elif user_type == "admin" or user_type == "poc":
        acc_list = user_details.objects.filter(email_id=user_name).values('account_name').distinct()
        for acc in acc_list:
            set = {}
            set['account_name'] = acc['account_name']
            accounts.append(set)

    else:
        pass

    return accounts
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LOAD USER RELATED PROJECTS LIST ----------------------------------------------------
def user_project_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        account_name = request.POST.get('account_name')
        user_type = request.user.email

        data['proj_name'] = user_proj_list(user_type, request.user.username, account_name)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)

def user_proj_list(user_type, user_name, account_name):
    projects = []
    if user_type == "superuser" or user_type == "leader" or user_type == "admin":
        proj_list = projects_list.objects.filter(account_name=account_name).values('project_name')
        for proj in proj_list:
            set = {}
            set['project_name'] = proj['project_name']
            projects.append(set)

    elif user_type == "poc":
        proj_list = user_details.objects.filter(email_id=user_name,
                                                account_name=account_name).values('project_name').distinct()
        for proj in proj_list:
            set = {}
            set['project_name'] = proj['project_name']
            projects.append(set)

    else:
        pass

    return projects
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- BUSINESS VERTICAL And USER STATISTICS ----------------------------------------------
def business_statistics(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_type = request.user.email

        t_updates = 0
        p_updates = 0
        ack_updates = 0
        t_kc = 0
        p_kc = 0
        ack_kc = 0

        t_comp_kc_percentage = 0.0
        t_recipients = 0
        t_pass_recipients = 0

        acc_list = []
        acc_t_updates = []
        acc_p_updates = []
        acc_ack_updates = []
        acc_t_kc = []
        acc_p_kc = []
        acc_ack_kc = []

        proj_list = []
        proj_t_updates = []
        proj_p_updates = []
        proj_ack_updates = []
        proj_t_kc = []
        proj_p_kc = []
        proj_ack_kc = []

        sp_list = []
        sp_t_updates = []
        sp_p_updates = []
        sp_ack_updates = []
        sp_t_kc = []
        sp_p_kc = []
        sp_ack_kc = []

        up_source_list = []
        up_source_t_updates = []
        up_source_p_updates = []
        up_source_ack_updates = []

        up_type_list = []
        up_type_t_updates = []
        up_type_p_updates = []
        up_type_ack_updates = []

        total_recipients_up = 0
        total_yes_up = 0

        acc = user_acc_list(user_type, request.user.username)

        if acc != []:
            for i in acc:
                t_acc_up = 0
                p_acc_up = 0
                ack_acc_up = 0
                t_acc_kc = 0
                p_acc_kc = 0
                ack_acc_kc = 0
                acc_list.append(i['account_name'])

                proj = user_proj_list(user_type, request.user.username, i['account_name'])

                if proj != []:
                    for j in proj:
                        t_proj_up = 0
                        p_proj_up = 0
                        ack_proj_up = 0
                        t_proj_kc = 0
                        p_proj_kc = 0
                        ack_proj_kc = 0
                        proj_list.append(j['project_name'])

                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')

                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                sp_list.append(k['sub_process_name'])

                                t_sp_up = 0
                                ack_sp_up = 0
                                t_sp_kc = 0
                                ack_sp_kc = 0

                                update_list = update_details.objects.filter(sub_process=k['sub_process_name'])
                                for up in update_list:
                                    total_r = update_recipients.objects.filter(update_id=up.id).count()
                                    ack_r = update_recipients.objects.filter(update_id=up.id).exclude(
                                        ack_status="NA").count()
                                    ack_aware_r = update_recipients.objects.filter(update_id=up.id,
                                                                                   ack_status="yes").count()

                                    total_recipients_up = total_recipients_up + total_r
                                    total_yes_up = total_yes_up + ack_aware_r

                                    t_sp_up = t_sp_up + 1
                                    if ack_r == 0:
                                        c_rate = 0.0
                                    else:
                                        c_rate = round(float((ack_r / total_r) * 100), 2)

                                    if c_rate == 100.0:
                                        ack_sp_up = ack_sp_up + 1

                                p_sp_up = t_sp_up - ack_sp_up

                                kc_list = conquest_list.objects.filter(sub_process=k['sub_process_name'])
                                for kc in kc_list:
                                    total_r = conquest_recipients.objects.filter(conquest_id=kc.id).count()
                                    ack_r = conquest_recipients.objects.filter(conquest_id=kc.id).exclude(
                                        attempts="0").count()
                                    t_recipients = t_recipients + total_r

                                    t_sp_kc = t_sp_kc + 1

                                    if ack_r == 0:
                                        c_rate = 0.0
                                    else:
                                        c_rate = round(float((ack_r / total_r) * 100), 2)

                                    t_comp_kc_percentage = t_comp_kc_percentage + c_rate

                                    if c_rate == 100.0:
                                        ack_sp_kc = ack_sp_kc + 1

                                    recipients = conquest_recipients.objects.filter(conquest_id=kc.id,
                                                                                    status='checked').values(
                                        'score').exclude(attempts=0)
                                    for kc_score in recipients:
                                        if float(kc_score['score']) >= float(kc.passing_score):
                                            t_pass_recipients = t_pass_recipients + 1

                                p_sp_kc = t_sp_kc - ack_sp_kc

                                sp_t_updates.append(t_sp_up)
                                sp_p_updates.append(p_sp_up)
                                sp_ack_updates.append(ack_sp_up)
                                sp_t_kc.append(t_sp_kc)
                                sp_p_kc.append(p_sp_kc)
                                sp_ack_kc.append(ack_sp_kc)

                                t_proj_up = t_proj_up + t_sp_up
                                p_proj_up = p_proj_up + p_sp_up
                                ack_proj_up = ack_proj_up + ack_sp_up
                                t_proj_kc = t_proj_kc + t_sp_kc
                                p_proj_kc = p_proj_kc + p_sp_kc
                                ack_proj_kc = ack_proj_kc + ack_sp_kc

                        proj_t_updates.append(t_proj_up)
                        proj_p_updates.append(p_proj_up)
                        proj_ack_updates.append(ack_proj_up)
                        proj_t_kc.append(t_proj_kc)
                        proj_p_kc.append(p_proj_kc)
                        proj_ack_kc.append(ack_proj_kc)

                        t_acc_up = t_acc_up + t_proj_up
                        p_acc_up = p_acc_up + p_proj_up
                        ack_acc_up = ack_acc_up + ack_proj_up
                        t_acc_kc = t_acc_kc + t_proj_kc
                        p_acc_kc = p_acc_kc + p_proj_kc
                        ack_acc_kc = ack_acc_kc + ack_proj_kc

                acc_t_updates.append(t_acc_up)
                acc_p_updates.append(p_acc_up)
                acc_ack_updates.append(ack_acc_up)
                acc_t_kc.append(t_acc_kc)
                acc_p_kc.append(p_acc_kc)
                acc_ack_kc.append(ack_acc_kc)

                t_updates = t_updates + t_acc_up
                p_updates = p_updates + p_acc_up
                ack_updates = ack_updates + ack_acc_up
                t_kc = t_kc + t_acc_kc
                p_kc = p_kc + p_acc_kc
                ack_kc = ack_kc + ack_acc_kc

        source_list = update_source_list.objects.all().values('update_source_name')
        for s in source_list:
            up_source_list.append(s['update_source_name'])
            source_t_up = 0
            source_ack_up = 0

            update_list = update_details.objects.filter(update_source=s['update_source_name'])
            for up in update_list:
                total_r = update_recipients.objects.filter(update_id=up.id).count()
                ack_r = update_recipients.objects.filter(update_id=up.id).exclude(ack_status="NA").count()
                source_t_up = source_t_up + 1

                if ack_r == 0:
                    c_rate = 0.0
                else:
                    c_rate = round(float((ack_r / total_r) * 100), 2)

                if c_rate == 100.0:
                    source_ack_up = source_ack_up + 1

            source_p_up = source_t_up - source_ack_up

            up_source_t_updates.append(source_t_up)
            up_source_p_updates.append(source_p_up)
            up_source_ack_updates.append(source_ack_up)

        type_list = update_type_list.objects.all().values('update_type_name')
        for t in type_list:
            up_type_list.append(t['update_type_name'])

            type_t_up = 0
            type_ack_up = 0

            update_list = update_details.objects.filter(update_type=t['update_type_name'])
            for up in update_list:
                total_r = update_recipients.objects.filter(update_id=up.id).count()
                ack_r = update_recipients.objects.filter(update_id=up.id).exclude(ack_status="NA").count()
                type_t_up = type_t_up + 1

                if ack_r == 0:
                    c_rate = 0.0
                else:
                    c_rate = round(float((ack_r / total_r) * 100), 2)

                if c_rate == 100.0:
                    type_ack_up = type_ack_up + 1

            type_p_up = type_t_up - type_ack_up

            up_type_t_updates.append(type_t_up)
            up_type_p_updates.append(type_p_up)
            up_type_ack_updates.append(type_ack_up)

        user_t_up_count = 0
        user_p_up_count = 0
        user_ack_up_count = 0
        user_exp_up_count = 0
        user_up_awareness_count = 0

        user_t_kc_count = 0
        user_p_kc_count = 0
        user_ack_kc_count = 0
        user_exp_kc_count = 0
        user_kc_t_score = 0.0

        user_pending_kc_count = 0

        user_kc_list = []
        user_kc_pass_percentage_list = []
        user_kc_score_percentage_list = []

        user_up_id_list = update_recipients.objects.filter(user_email=request.user.username)
        for user_up_id in user_up_id_list:
            user_t_up_count = user_t_up_count + 1

            if user_up_id.ack_status == "yes":
                user_ack_up_count = user_ack_up_count + 1
                user_up_awareness_count = user_up_awareness_count + 1

            elif user_up_id.ack_status == "no":
                user_ack_up_count = user_ack_up_count + 1

            else:
                up_data = update_details.objects.filter(id=user_up_id.update_id)
                current_timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')
                val = datetime.datetime.strptime(current_timestamp, "%d-%m-%Y %H:%M") < datetime.datetime.strptime(
                    up_data[0].deadline, "%d-%m-%Y %H:%M")
                if val:
                    user_p_up_count = user_p_up_count + 1
                else:
                    user_exp_up_count = user_exp_up_count + 1

        user_kc_id_list = conquest_recipients.objects.filter(user_email=request.user.username)
        for user_kc_id in user_kc_id_list:
            user_t_kc_count = user_t_kc_count + 1

            kc_data = conquest_list.objects.filter(id=user_kc_id.conquest_id)
            if user_kc_id.attempts != "0":
                user_ack_kc_count = user_ack_kc_count + 1
                user_kc_t_score = user_kc_t_score + float(user_kc_id.score)

                user_kc_list.append("KC_ID-" + str(user_kc_id.conquest_id))
                user_kc_pass_percentage_list.append(float(kc_data[0].passing_score))
                user_kc_score_percentage_list.append(float(user_kc_id.score))

            else:
                current_timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')
                val = datetime.datetime.strptime(current_timestamp, "%d-%m-%Y %H:%M") < datetime.datetime.strptime(
                    kc_data[0].deadline, "%d-%m-%Y %H:%M")

                if val:
                    user_p_kc_count = user_p_kc_count + 1

                    up_kc_count = update_conquest.objects.filter(conquest_id=user_kc_id.conquest_id).count()
                    if up_kc_count == 0:
                        user_pending_kc_count = user_pending_kc_count + 1
                else:
                    user_exp_kc_count = user_exp_kc_count + 1

        data['user_pending_up_count'] = user_p_up_count
        data['user_pending_kc_count'] = user_pending_kc_count

        data['user_t_up_count'] = user_t_up_count
        data['user_p_up_count'] = user_p_up_count
        data['user_ack_up_count'] = user_ack_up_count
        data['user_exp_up_count'] = user_exp_up_count
        if user_up_awareness_count == 0 or user_ack_up_count == 0:
            data['user_up_awareness_percentage'] = 0.0
        else:
            data['user_up_awareness_percentage'] = (float(user_up_awareness_count) / float(user_ack_up_count)) * 100

        data['user_up_stats'] = []
        data['user_up_stats'].append(data['user_t_up_count'])
        data['user_up_stats'].append(data['user_p_up_count'])
        data['user_up_stats'].append(data['user_ack_up_count'])
        data['user_up_stats'].append(data['user_exp_up_count'])

        data['user_t_kc_count'] = user_t_kc_count
        data['user_p_kc_count'] = user_p_kc_count
        data['user_ack_kc_count'] = user_ack_kc_count
        data['user_exp_kc_count'] = user_exp_kc_count
        if user_kc_t_score == 0.0 or user_ack_kc_count == 0:
            data['user_kc_avg_score'] = 0.0
        else:
            data['user_kc_avg_score'] = user_kc_t_score / float(user_ack_kc_count)

        data['user_kc_stats'] = []
        data['user_kc_stats'].append(data['user_t_kc_count'])
        data['user_kc_stats'].append(data['user_p_kc_count'])
        data['user_kc_stats'].append(data['user_ack_kc_count'])
        data['user_kc_stats'].append(data['user_exp_kc_count'])

        data['user_kc_list'] = user_kc_list
        data['user_kc_pass_percentage_list'] = user_kc_pass_percentage_list
        data['user_kc_score_percentage_list'] = user_kc_score_percentage_list

        data['t_updates'] = t_updates
        data['p_updates'] = p_updates
        data['ack_updates'] = ack_updates

        if total_yes_up == 0 or total_recipients_up == 0:
            data['updates_awareness_percentage'] = 0.0
        else:
            data['updates_awareness_percentage'] = (float(total_yes_up) / float(total_recipients_up)) * 100

        data['t_kc'] = t_kc
        data['p_kc'] = p_kc
        data['ack_kc'] = ack_kc

        if t_kc == 0 or t_comp_kc_percentage == 0.0:
            data['kc_completion_rate'] = 0.0
        else:
            data['kc_completion_rate'] = t_comp_kc_percentage / float(t_kc)

        if t_pass_recipients == 0 or t_recipients == 0:
            data['kc_passing_rate'] = 0.0
        else:
            data['kc_passing_rate'] = (float(t_pass_recipients) / float(t_recipients)) * 100

        data['acc_list'] = acc_list
        data['acc_t_updates'] = acc_t_updates
        data['acc_p_updates'] = acc_p_updates
        data['acc_ack_updates'] = acc_ack_updates
        data['acc_t_kc'] = acc_t_kc
        data['acc_p_kc'] = acc_p_kc
        data['acc_ack_kc'] = acc_ack_kc

        data['proj_list'] = proj_list
        data['proj_t_updates'] = proj_t_updates
        data['proj_p_updates'] = proj_p_updates
        data['proj_ack_updates'] = proj_ack_updates
        data['proj_t_kc'] = proj_t_kc
        data['proj_p_kc'] = proj_p_kc
        data['proj_ack_kc'] = proj_ack_kc

        data['sp_list'] = sp_list
        data['sp_t_updates'] = sp_t_updates
        data['sp_p_updates'] = sp_p_updates
        data['sp_ack_updates'] = sp_ack_updates
        data['sp_t_kc'] = sp_t_kc
        data['sp_p_kc'] = sp_p_kc
        data['sp_ack_kc'] = sp_ack_kc

        data['update_source_list'] = up_source_list
        data['update_source_t_updates'] = up_source_t_updates
        data['update_source_p_updates'] = up_source_p_updates
        data['update_source_ack_updates'] = up_source_ack_updates

        data['update_type_list'] = up_type_list
        data['update_type_t_updates'] = up_type_t_updates
        data['update_type_p_updates'] = up_type_p_updates
        data['update_type_ack_updates'] = up_type_ack_updates

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- VERTICAL STATISTICS ON DATE PICKUP -------------------------------------------------
def dashboard_vertical_data_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        user_type = request.user.email

        t_updates = 0
        p_updates = 0
        ack_updates = 0
        t_kc = 0
        p_kc = 0
        ack_kc = 0

        t_comp_kc_percentage = 0.0
        t_recipients = 0
        t_pass_recipients = 0

        acc_list = []
        acc_t_updates = []
        acc_p_updates = []
        acc_ack_updates = []
        acc_t_kc = []
        acc_p_kc = []
        acc_ack_kc = []

        proj_list = []
        proj_t_updates = []
        proj_p_updates = []
        proj_ack_updates = []
        proj_t_kc = []
        proj_p_kc = []
        proj_ack_kc = []

        sp_list = []
        sp_t_updates = []
        sp_p_updates = []
        sp_ack_updates = []
        sp_t_kc = []
        sp_p_kc = []
        sp_ack_kc = []

        up_source_list = []
        up_source_t_updates = []
        up_source_p_updates = []
        up_source_ack_updates = []

        up_type_list = []
        up_type_t_updates = []
        up_type_p_updates = []
        up_type_ack_updates = []

        total_recipients_up = 0
        total_yes_up = 0

        acc = user_acc_list(user_type, request.user.username)

        if acc != []:
            for i in acc:
                t_acc_up = 0
                p_acc_up = 0
                ack_acc_up = 0
                t_acc_kc = 0
                p_acc_kc = 0
                ack_acc_kc = 0
                acc_list.append(i['account_name'])

                proj = user_proj_list(user_type, request.user.username, i['account_name'])

                if proj != []:
                    for j in proj:
                        t_proj_up = 0
                        p_proj_up = 0
                        ack_proj_up = 0
                        t_proj_kc = 0
                        p_proj_kc = 0
                        ack_proj_kc = 0
                        proj_list.append(j['project_name'])

                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')

                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                t_sp_up = 0
                                p_sp_up = 0
                                ack_sp_up = 0

                                t_sp_kc = 0
                                p_sp_kc = 0
                                ack_sp_kc = 0

                                sp_list.append(k['sub_process_name'])

                                up_sp_list = update_details.objects.filter(sub_process=k['sub_process_name'])
                                for up_sp_id in up_sp_list:
                                    val1 = datetime.datetime.strptime(up_sp_id.timestamp[0:10],
                                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                                        start_date, "%d-%m-%Y")
                                    val2 = datetime.datetime.strptime(up_sp_id.timestamp[0:10],
                                                                      "%d-%m-%Y") <= datetime.datetime.strptime(
                                        end_date, "%d-%m-%Y")

                                    if val1 and val2:
                                        total_r = update_recipients.objects.filter(update_id=up_sp_id.id).count()
                                        ack_r = update_recipients.objects.filter(update_id=up_sp_id.id).exclude(
                                            ack_status="NA").count()
                                        ack_aware_r = update_recipients.objects.filter(update_id=up_sp_id.id,
                                                                                       ack_status="yes").count()

                                        t_sp_up = t_sp_up + 1
                                        total_recipients_up = total_recipients_up + total_r
                                        total_yes_up = total_yes_up + ack_aware_r

                                        if ack_r == 0:
                                            c_rate = 0.0
                                        else:
                                            c_rate = round(float((ack_r / total_r) * 100), 2)

                                        if c_rate == 100.0:
                                            ack_sp_up = ack_sp_up + 1
                                        else:
                                            p_sp_up = p_sp_up + 1

                                kc_sp_list = conquest_list.objects.filter(sub_process=k['sub_process_name'])
                                for kc_sp_id in kc_sp_list:

                                    val1 = datetime.datetime.strptime(kc_sp_id.timestamp[0:10],
                                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                                        start_date, "%d-%m-%Y")
                                    val2 = datetime.datetime.strptime(kc_sp_id.timestamp[0:10],
                                                                      "%d-%m-%Y") <= datetime.datetime.strptime(
                                        end_date, "%d-%m-%Y")

                                    if val1 and val2:
                                        total_r = conquest_recipients.objects.filter(conquest_id=kc_sp_id.id).count()
                                        ack_r = conquest_recipients.objects.filter(conquest_id=kc_sp_id.id).exclude(
                                            attempts="0").count()
                                        t_recipients = t_recipients + total_r

                                        t_sp_kc = t_sp_kc + 1

                                        if ack_r == 0:
                                            c_rate = 0.0
                                        else:
                                            c_rate = round(float((ack_r / total_r) * 100), 2)

                                        t_comp_kc_percentage = t_comp_kc_percentage + c_rate

                                        if c_rate == 100.0:
                                            ack_sp_kc = ack_sp_kc + 1

                                        recipients = conquest_recipients.objects.filter(conquest_id=kc_sp_id.id,
                                                                                        status='checked').values(
                                            'score').exclude(attempts=0)
                                        for kc_score in recipients:
                                            if kc_score['score'] >= kc_sp_id.passing_score:
                                                t_pass_recipients = t_pass_recipients + 1

                                        p_sp_kc = t_sp_kc - ack_sp_kc

                                sp_t_updates.append(t_sp_up)
                                sp_p_updates.append(p_sp_up)
                                sp_ack_updates.append(ack_sp_up)
                                sp_t_kc.append(t_sp_kc)
                                sp_p_kc.append(p_sp_kc)
                                sp_ack_kc.append(ack_sp_kc)

                                t_proj_up = t_proj_up + t_sp_up
                                p_proj_up = p_proj_up + p_sp_up
                                ack_proj_up = ack_proj_up + ack_sp_up
                                t_proj_kc = t_proj_kc + t_sp_kc
                                p_proj_kc = p_proj_kc + p_sp_kc
                                ack_proj_kc = ack_proj_kc + ack_sp_kc

                        proj_t_updates.append(t_proj_up)
                        proj_p_updates.append(p_proj_up)
                        proj_ack_updates.append(ack_proj_up)
                        proj_t_kc.append(t_proj_kc)
                        proj_p_kc.append(p_proj_kc)
                        proj_ack_kc.append(ack_proj_kc)

                        t_acc_up = t_acc_up + t_proj_up
                        p_acc_up = p_acc_up + p_proj_up
                        ack_acc_up = ack_acc_up + ack_proj_up
                        t_acc_kc = t_acc_kc + t_proj_kc
                        p_acc_kc = p_acc_kc + p_proj_kc
                        ack_acc_kc = ack_acc_kc + ack_proj_kc

                acc_t_updates.append(t_acc_up)
                acc_p_updates.append(p_acc_up)
                acc_ack_updates.append(ack_acc_up)
                acc_t_kc.append(t_acc_kc)
                acc_p_kc.append(p_acc_kc)
                acc_ack_kc.append(ack_acc_kc)

                t_updates = t_updates + t_acc_up
                p_updates = p_updates + p_acc_up
                ack_updates = ack_updates + ack_acc_up
                t_kc = t_kc + t_acc_kc
                p_kc = p_kc + p_acc_kc
                ack_kc = ack_kc + ack_acc_kc

        source_list = update_source_list.objects.all().values('update_source_name')
        for s in source_list:
            source_t_up = 0
            source_p_up = 0
            source_ack_up = 0
            up_source_list.append(s['update_source_name'])

            source_up_list = update_details.objects.filter(update_source=s['update_source_name'])
            for sp in source_up_list:
                val1 = datetime.datetime.strptime(sp.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                    start_date, "%d-%m-%Y")
                val2 = datetime.datetime.strptime(sp.timestamp[0:10],
                                                  "%d-%m-%Y") <= datetime.datetime.strptime(
                    end_date, "%d-%m-%Y")

                if val1 and val2:
                    total_r = update_recipients.objects.filter(update_id=sp.id).count()
                    ack_r = update_recipients.objects.filter(update_id=sp.id).exclude(ack_status="NA").count()
                    source_t_up = source_t_up + 1

                    if ack_r == 0:
                        c_rate = 0.0
                    else:
                        c_rate = round(float((ack_r / total_r) * 100), 2)

                    if c_rate == 100.0:
                        source_ack_up = source_ack_up + 1

                source_p_up = source_t_up - source_ack_up

            up_source_t_updates.append(source_t_up)
            up_source_p_updates.append(source_p_up)
            up_source_ack_updates.append(source_ack_up)

        type_list = update_type_list.objects.all().values('update_type_name')
        for t in type_list:
            type_t_up = 0
            type_p_up = 0
            type_ack_up = 0
            up_type_list.append(t['update_type_name'])

            type_up_list = update_details.objects.filter(update_type=t['update_type_name'])
            for type_up in type_up_list:
                val1 = datetime.datetime.strptime(type_up.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                    start_date, "%d-%m-%Y")
                val2 = datetime.datetime.strptime(type_up.timestamp[0:10],
                                                  "%d-%m-%Y") <= datetime.datetime.strptime(
                    end_date, "%d-%m-%Y")

                if val1 and val2:
                    total_r = update_recipients.objects.filter(update_id=type_up.id).count()
                    ack_r = update_recipients.objects.filter(update_id=type_up.id).exclude(ack_status="NA").count()
                    type_t_up = type_t_up + 1

                    if ack_r == 0:
                        c_rate = 0.0
                    else:
                        c_rate = round(float((ack_r / total_r) * 100), 2)

                    if c_rate == 100.0:
                        type_ack_up = type_ack_up + 1

                type_p_up = type_t_up - type_ack_up

            up_type_t_updates.append(type_t_up)
            up_type_p_updates.append(type_p_up)
            up_type_ack_updates.append(type_ack_up)

        data['t_updates'] = t_updates
        data['p_updates'] = p_updates
        data['ack_updates'] = ack_updates

        if total_yes_up == 0 or total_recipients_up == 0:
            data['updates_awareness_percentage'] = 0.0
        else:
            data['updates_awareness_percentage'] = (float(total_yes_up) / float(total_recipients_up)) * 100

        data['t_kc'] = t_kc
        data['p_kc'] = p_kc
        data['ack_kc'] = ack_kc

        if t_kc == 0 or t_comp_kc_percentage == 0.0:
            data['kc_completion_rate'] = 0.0
        else:
            data['kc_completion_rate'] = t_comp_kc_percentage / float(t_kc)

        if t_pass_recipients == 0 or t_recipients == 0:
            data['kc_passing_rate'] = 0.0
        else:
            data['kc_passing_rate'] = (float(t_pass_recipients) / float(t_recipients)) * 100

        data['acc_list'] = acc_list
        data['acc_t_updates'] = acc_t_updates
        data['acc_p_updates'] = acc_p_updates
        data['acc_ack_updates'] = acc_ack_updates
        data['acc_t_kc'] = acc_t_kc
        data['acc_p_kc'] = acc_p_kc
        data['acc_ack_kc'] = acc_ack_kc

        data['proj_list'] = proj_list
        data['proj_t_updates'] = proj_t_updates
        data['proj_p_updates'] = proj_p_updates
        data['proj_ack_updates'] = proj_ack_updates
        data['proj_t_kc'] = proj_t_kc
        data['proj_p_kc'] = proj_p_kc
        data['proj_ack_kc'] = proj_ack_kc

        data['sp_list'] = sp_list
        data['sp_t_updates'] = sp_t_updates
        data['sp_p_updates'] = sp_p_updates
        data['sp_ack_updates'] = sp_ack_updates
        data['sp_t_kc'] = sp_t_kc
        data['sp_p_kc'] = sp_p_kc
        data['sp_ack_kc'] = sp_ack_kc

        data['update_source_list'] = up_source_list
        data['update_source_t_updates'] = up_source_t_updates
        data['update_source_p_updates'] = up_source_p_updates
        data['update_source_ack_updates'] = up_source_ack_updates

        data['update_type_list'] = up_type_list
        data['update_type_t_updates'] = up_type_t_updates
        data['update_type_p_updates'] = up_type_p_updates
        data['update_type_ack_updates'] = up_type_ack_updates

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- USER STATISTICS ON DATE PICKUP -----------------------------------------------------
def dashboard_user_data_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        user_t_up_count = 0
        user_p_up_count = 0
        user_ack_up_count = 0
        user_exp_up_count = 0
        user_up_awareness_count = 0

        user_t_kc_count = 0
        user_p_kc_count = 0
        user_ack_kc_count = 0
        user_exp_kc_count = 0
        user_kc_t_score = 0.0

        user_kc_list = []
        user_kc_pass_percentage_list = []
        user_kc_score_percentage_list = []

        user_up_id_list = update_recipients.objects.filter(user_email=request.user.username)
        for user_up_id in user_up_id_list:
            up_data = update_details.objects.filter(id=user_up_id.update_id)
            val1 = datetime.datetime.strptime(up_data[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(up_data[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                end_date, "%d-%m-%Y")

            if val1 and val2:
                user_t_up_count = user_t_up_count + 1

                if user_up_id.ack_status == "yes":
                    user_ack_up_count = user_ack_up_count + 1
                    user_up_awareness_count = user_up_awareness_count + 1

                elif user_up_id.ack_status == "no":
                    user_ack_up_count = user_ack_up_count + 1

                else:
                    current_timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')
                    val = datetime.datetime.strptime(current_timestamp, "%d-%m-%Y %H:%M") < datetime.datetime.strptime(
                        up_data[0].deadline, "%d-%m-%Y %H:%M")

                    if val:
                        user_p_up_count = user_p_up_count + 1
                    else:
                        user_exp_up_count = user_exp_up_count + 1

        user_kc_id_list = conquest_recipients.objects.filter(user_email=request.user.username)
        for user_kc_id in user_kc_id_list:
            kc_data = conquest_list.objects.filter(id=user_kc_id.conquest_id)
            val1 = datetime.datetime.strptime(kc_data[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(kc_data[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                end_date, "%d-%m-%Y")

            if val1 and val2:
                user_t_kc_count = user_t_kc_count + 1

                if user_kc_id.attempts != "0":
                    user_ack_kc_count = user_ack_kc_count + 1
                    user_kc_t_score = user_kc_t_score + float(user_kc_id.score)

                    user_kc_list.append("KC_ID-" + str(user_kc_id.conquest_id))
                    user_kc_pass_percentage_list.append(float(kc_data[0].passing_score))
                    user_kc_score_percentage_list.append(float(user_kc_id.score))

                else:
                    current_timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')
                    val = datetime.datetime.strptime(current_timestamp, "%d-%m-%Y %H:%M") < datetime.datetime.strptime(
                        kc_data[0].deadline, "%d-%m-%Y %H:%M")

                    if val:
                        user_p_kc_count = user_p_kc_count + 1
                    else:
                        user_exp_kc_count = user_exp_kc_count + 1

        data['user_t_up_count'] = user_t_up_count
        data['user_p_up_count'] = user_p_up_count
        data['user_ack_up_count'] = user_ack_up_count
        data['user_exp_up_count'] = user_exp_up_count
        if user_up_awareness_count == 0 or user_ack_up_count == 0:
            data['user_up_awareness_percentage'] = 0.0
        else:
            data['user_up_awareness_percentage'] = (float(user_up_awareness_count) / float(user_ack_up_count)) * 100

        data['user_up_stats'] = []
        data['user_up_stats'].append(data['user_t_up_count'])
        data['user_up_stats'].append(data['user_p_up_count'])
        data['user_up_stats'].append(data['user_ack_up_count'])
        data['user_up_stats'].append(data['user_exp_up_count'])

        data['user_t_kc_count'] = user_t_kc_count
        data['user_p_kc_count'] = user_p_kc_count
        data['user_ack_kc_count'] = user_ack_kc_count
        data['user_exp_kc_count'] = user_exp_kc_count
        if user_kc_t_score == 0.0 or user_ack_kc_count == 0:
            data['user_kc_avg_score'] = 0.0
        else:
            data['user_kc_avg_score'] = user_kc_t_score / float(user_ack_kc_count)

        data['user_kc_stats'] = []
        data['user_kc_stats'].append(data['user_t_kc_count'])
        data['user_kc_stats'].append(data['user_p_kc_count'])
        data['user_kc_stats'].append(data['user_ack_kc_count'])
        data['user_kc_stats'].append(data['user_exp_kc_count'])

        data['user_kc_list'] = user_kc_list
        data['user_kc_pass_percentage_list'] = user_kc_pass_percentage_list
        data['user_kc_score_percentage_list'] = user_kc_score_percentage_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END DASHBOARD-----------------------------------------------------------------------





# --------------------------------- REPORTING --------------------------------------------------------------------------


# --------------------------------- REDIRECT TO DASHBOARD --------------------------------------------------------------
@login_required
def reporting(request):
    request.session.set_expiry(3600)
    return render(request, 'b2-reporting.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ACCOUNT/PROJECT ON DATE PICKUP STATISTICS ------------------------------------------
def reporting_data_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        proj_name = request.POST.get('project_name')

        emp_data = []

        emp_list = user_details.objects.filter(project_name=proj_name).values('emp_id', 'email_id').order_by(
            'emp_id').distinct()

        for x in emp_list:

            user_type = User.objects.filter(username = x['email_id']).values('email')

            if user_type[0]['email'] == "superuser" or user_type[0]['email'] == "agent":
                set = {}
                updates_list = update_recipients.objects.filter(user_email=x['email_id'])
                kc_list = conquest_recipients.objects.filter(user_email=x['email_id'])
                t_updates = 0
                ack_updates = 0
                pending_updates = 0

                t_kc = 0
                completed_kc = 0
                pending_kc = 0

                for y in updates_list:
                    upd = update_details.objects.filter(pk=y.update_id)
                    val1 = datetime.datetime.strptime(upd[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date, "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(upd[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                        end_date, "%d-%m-%Y")

                    if val1 and val2:
                        t_updates = t_updates + 1
                        if y.ack_status == "NA":
                            pending_updates = pending_updates + 1
                        else:
                            ack_updates = ack_updates + 1

                for y in kc_list:
                    kcd = conquest_list.objects.filter(pk=y.conquest_id)
                    val1 = datetime.datetime.strptime(kcd[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date, "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(kcd[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                        end_date, "%d-%m-%Y")

                    if val1 and val2:
                        t_kc = t_kc + 1
                        if y.attempts == "0":
                            pending_kc = pending_kc + 1
                        else:
                            completed_kc = completed_kc + 1

                emp_f_name = User.objects.filter(username = x['email_id']).values('first_name')[0]['first_name']
                emp_l_name = User.objects.filter(username = x['email_id']).values('last_name')[0]['last_name']
                emp_name = emp_f_name + " " + emp_l_name

                set['emp_id']               = x['emp_id']
                set['emp_email']            = x['email_id']
                set['emp_name']             = emp_name
                set['total_updates']        = t_updates
                set['acknowledged_updates'] = ack_updates
                set['pending_updates']      = pending_updates
                set['total_kc']             = t_kc
                set['completed_kc']         = completed_kc
                set['pending_kc']           = pending_kc
                emp_data.append(set)

        data['employee_details'] = emp_data

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST OF Updates On EMPLOYEE SELECTION ----------------------------------------------
def reporting_emp_update_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        emp_email = request.POST.get('emp_email')

        emp_update_list = []
        update_list = update_recipients.objects.filter(user_email=emp_email).values('update_id', 'timestamp', 'ack_status', 'ack_comment')

        for x in update_list:
            set = {}
            upd = update_details.objects.filter(pk=x['update_id'])
            val1 = datetime.datetime.strptime(upd[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(upd[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                end_date, "%d-%m-%Y")

            if val1 and val2:
                set['update_id']            = x['update_id']
                set['timestamp']            = x['timestamp']
                set['update_title']         = upd[0].title
                set['acknowledged_status']  = x['ack_status']
                set['acknowledged_comment'] = x['ack_comment']
                emp_update_list.append(set)

        data['emp_update_list'] = emp_update_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST OF Knowledge Check On EMPLOYEE SELECTION --------------------------------------
def reporting_emp_kc_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        emp_email = request.POST.get('emp_email')

        emp_kc_list = []
        kc_list = conquest_recipients.objects.filter(user_email=emp_email).values('conquest_id', 'timestamp', 'attempts', 'score')

        for x in kc_list:
            set = {}
            kcd = conquest_list.objects.filter(pk=x['conquest_id'])
            val1 = datetime.datetime.strptime(kcd[0].timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(kcd[0].timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                end_date, "%d-%m-%Y")

            ps = conquest_list.objects.filter(pk=x['conquest_id']).values('passing_score', 'title')
            passing_score = ps[0]['passing_score']

            if x['attempts'] != "0":
                if float(x['score']) < float(passing_score):
                    ack_status = "FAILED"
                else:
                    ack_status = "PASSED"
            else:
                ack_status = "NA"

            if val1 and val2:
                set['kc_id']        = x['conquest_id']
                set['kc_timestamp'] = x['timestamp']
                set['kc_title']     = ps[0]['title']
                set['kc_attempts']  = x['attempts']
                set['kc_score']     = round(float(x['score']), 2)
                set['kc_status']    = ack_status
                emp_kc_list.append(set)

        data['emp_kc_list'] = emp_kc_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END REPORTING-----------------------------------------------------------------------





# --------------------------------- My Tasks - Updates + Knowledge Check List ------------------------------------------


# --------------------------------- REDIRECT TO My Tasks - Updates + Knowledge Check List ------------------------------
@login_required
def mytasks_up_kc(request):
    request.session.set_expiry(3600)
    update_id_list = update_recipients.objects.filter(user_email=request.user.username).annotate(
        update_id_field=Cast('update_id', IntegerField())).order_by('-update_id_field')

    updates_list = []

    for x in update_id_list:
        upd = update_details.objects.filter(pk=x.update_id)
        ack_list = update_recipients.objects.filter(user_email=request.user.username, update_id=x.update_id)

        set = {}
        set['timestamp']    = upd[0].timestamp
        set['title']        = upd[0].title
        set['update_id']    = ack_list[0].update_id
        set['ack_status']   = ack_list[0].ack_status
        updates_list.append(set)

    data = {}
    data['updates_details'] = updates_list

    return render(request, 'b3a-mytasks-updates-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- My Tasks - Updates + Knowledge Check List as per DATE Selected ---------------------
def mytask_up_kc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        update_id_list = update_recipients.objects.filter(user_email=request.user.username).annotate(
            update_id_field=Cast('update_id', IntegerField())).order_by('-update_id_field')

        updates_list = []
        for x in update_id_list:
            upd = update_details.objects.filter(pk=x.update_id)
            for y in upd:
                val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                    start_date, "%d-%m-%Y")
                val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                               "%d-%m-%Y")
                if val1:
                    if val2:
                        ack_list = update_recipients.objects.filter(user_email=request.user.username,
                                                                    update_id=x.update_id)

                        set = {}
                        set['timestamp']        = upd[0].timestamp
                        set['title']            = upd[0].title
                        set['update_id']        = ack_list[0].update_id
                        set['ack_status']       = ack_list[0].ack_status
                        updates_list.append(set)

        data['update_details'] = updates_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE ACKNOWLEDGEMENT AND ANSWERS OF KNOWLEDGE TEST GIVEN BY EMPLOYEE ---------------
def save_user_ack(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        update_id = request.POST.get('update_id')
        ack_status = request.POST.get('ack_status')
        ack_comment = request.POST.get('ack_comment')
        up_sender = request.POST.get('sender')
        up_reciever = request.user.username
        answers = json.loads(request.POST['answers'])
        timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        agent_email = request.user.username

        update_recipients.objects.filter(update_id=update_id, user_email=request.user.username) \
            .update(ack_status=ack_status, ack_comment=ack_comment, timestamp = timestamp)

        if ack_comment == "NA":
            comments = ""
        else:
            comments = ack_comment

        answer_list = []

        subj = ""
        for check in answers:
            if check['question_type'] == "Subjective":
                subj = "yes"

        if ack_status == "yes":
            conquest_id_list = update_conquest.objects.filter(update_id=update_id).count()
            if conquest_id_list > 0:
                conquest_id = update_conquest.objects.filter(update_id=update_id)
                ques_id_list = conquest_questions.objects.filter(conquest_id=conquest_id[0].conquest_id)

                for q in ques_id_list:
                    ques_op = question_details.objects.filter(question_id=q.question_id)
                    count = 0
                    set = {}
                    answ = []
                    for q_op in ques_op:
                        count = count + 1
                        if q_op.question_answer == "yes":
                            answ.append(count)

                    set['question_id'] = q.question_id
                    set['question_answer'] = answ
                    answer_list.append(set)

                total_marks = 0
                max_marks = 0
                atmpts = conquest_recipients.objects.filter(conquest_id=conquest_id[0].conquest_id,
                                                            user_email=request.user.username)
                attempts = int(atmpts[0].attempts) + 1

                for i in answers:
                    save_user_ack_answers(conquest_id[0].conquest_id, i['question_id'], up_reciever, attempts,
                                          i['question_answer'])

                if subj != "yes":
                    for i in answers:
                        if i['question_type'] == "MAQs":
                            for j in answer_list:
                                if i['question_id'] == j['question_id']:
                                    q_max_marks         = int(i['question_marks'])
                                    t_q_op              = question_details.objects.filter(question_id=i['question_id']).count()
                                    t_q_correct_op      = len(j['question_answer'])
                                    t_u_op_selected     = len(i['question_answer'])
                                    t_u_correct_op      = 0
                                    t_u_incorrect_op    = 0
                                    if t_u_op_selected > t_q_correct_op:
                                        t_extra_op_selected = t_u_op_selected - t_q_correct_op
                                    else:
                                        t_extra_op_selected = 0

                                    if t_q_op != t_u_op_selected:
                                        for op_eval in range(t_u_op_selected):
                                            if i['question_answer'][op_eval] in j['question_answer']:
                                                t_u_correct_op = t_u_correct_op + 1
                                            else:
                                                t_u_incorrect_op = t_u_incorrect_op + 1

                                        user_q_marks = (((t_u_op_selected - t_u_incorrect_op) * q_max_marks) / t_q_correct_op) - (t_extra_op_selected / t_q_op)

                                        total_marks = total_marks + user_q_marks

                                    max_marks = max_marks + int(i['question_marks'])
                        else:
                            for j in answer_list:
                                if i['question_id'] == j['question_id']:
                                    if (i['question_answer'] == j['question_answer']):
                                        total_marks = total_marks + int(i['question_marks'])

                                    max_marks = max_marks + int(i['question_marks'])

                    u_score = (round(float(total_marks), 2) / float(max_marks)) * 100

                    conquest_recipients.objects \
                        .filter(conquest_id=conquest_id[0].conquest_id, user_email=request.user.username) \
                        .update(attempts=attempts, score=u_score, status="checked", timestamp = timestamp)

                else:
                    conquest_recipients.objects \
                        .filter(conquest_id=conquest_id[0].conquest_id, user_email=request.user.username) \
                        .update(attempts=attempts, status="pending", timestamp = timestamp)

                    send_knowledge_check_acknowledgement_email(up_sender, agent_email)

        send_acknowledgement_email(up_reciever, up_sender, ack_status, comments)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE ANSWERS OF KNOWLEDGE TEST -----------------------------------------------------
def save_user_ack_answers(conquest_id, question_id, user_email_id, user_attempt_no, answers):
    conq_user_ans = conquest_recipients_answers(conquest_id=conquest_id,
                                                question_id=question_id,
                                                user_email=user_email_id,
                                                attempt_no=user_attempt_no,
                                                answers=answers)
    conq_user_ans.save()
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END My Tasks - Updates + Knowledge Check List --------------------------------------





# --------------------------------- My Tasks - Knowledge Check List ----------------------------------------------------


# --------------------------------- REDIRECT TO My Tasks - Knowledge Check List ----------------------------------------
@login_required
def mytasks_kc(request):
    request.session.set_expiry(3600)
    kc_id_list = conquest_recipients.objects.filter(user_email=request.user.username).annotate(
        conquest_id_field=Cast('conquest_id', IntegerField())).order_by('-conquest_id_field')
    kc_list = []
    user_attempts = []

    for y in kc_id_list:
        kc_up_check = update_conquest.objects.filter(conquest_id=y.conquest_id).count()
        if kc_up_check == 0:
            kcd = conquest_list.objects.filter(pk=y.conquest_id)

            total_r = conquest_recipients.objects.filter(conquest_id=y.id).count()
            ack_r = conquest_recipients.objects.filter(conquest_id=y.id).exclude(
                attempts="0").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate = round(float((ack_r / total_r) * 100), 2)

            set = {}
            set['id'] = kcd[0].id
            set['title'] = kcd[0].title
            set['timestamp'] = kcd[0].timestamp
            set['sub_process'] = kcd[0].sub_process
            set['questions_count'] = kcd[0].questions_count
            set['passing_score'] = kcd[0].passing_score
            set['completion_rate'] = c_rate
            set['max_attempts'] = kcd[0].max_attempts
            set['sender'] = kcd[0].sender
            set['deadline'] = kcd[0].deadline
            kc_list.append(set)

            user_kc_details = conquest_recipients.objects.filter(user_email=request.user.username,
                                                                 conquest_id=y.conquest_id)

            set = {}
            set['conquest_id'] = user_kc_details[0].conquest_id
            set['user_email'] = user_kc_details[0].user_email
            set['attempts'] = user_kc_details[0].attempts
            set['score'] = user_kc_details[0].score
            set['status'] = user_kc_details[0].status
            user_attempts.append(set)

    data = {}
    data['conquest_details'] = kc_list
    data['conquest_attempts'] = user_attempts
    return render(request, 'b3b-mytasks-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- My Tasks - Knowledge Check List as per DATE Selected -------------------------------
def mytask_kc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        kc_id_list = conquest_recipients.objects.filter(user_email=request.user.username).order_by('-conquest_id')

        kc_list = []
        user_attempts = []
        for x in kc_id_list:
            kc_up_check = update_conquest.objects.filter(conquest_id=x.conquest_id).count()
            if kc_up_check == 0:
                kcd = conquest_list.objects.filter(pk=x.conquest_id)
                for y in kcd:
                    val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date, "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                        end_date, "%d-%m-%Y")
                    if val1:
                        if val2:

                            total_r = conquest_recipients.objects.filter(conquest_id = y.id).count()
                            ack_r = conquest_recipients.objects.filter(conquest_id = y.id).exclude(
                                attempts="0").count()

                            if ack_r == 0:
                                c_rate = 0.0
                            else:
                                c_rate = round(float((ack_r / total_r) * 100), 2)

                            set = {}
                            set['id']               = y.id
                            set['title']            = y.title
                            set['timestamp']        = y.timestamp
                            set['sub_process']      = y.sub_process
                            set['questions_count']  = y.questions_count
                            set['passing_score']    = y.passing_score
                            set['completion_rate']  = c_rate
                            set['max_attempts']     = y.max_attempts
                            set['sender']           = y.sender
                            set['deadline']         = y.deadline
                            kc_list.append(set)

                            user_kc_details = conquest_recipients.objects.filter(user_email=request.user.username, conquest_id=x.conquest_id)

                            set = {}
                            set['conquest_id']  = user_kc_details[0].conquest_id
                            set['user_email']   = user_kc_details[0].user_email
                            set['attempts']     = user_kc_details[0].attempts
                            set['score']        = user_kc_details[0].score
                            set['status']       = user_kc_details[0].status
                            user_attempts.append(set)

        data['conquest_details'] = kc_list
        data['conquest_user_details'] = user_attempts

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# ---------------------------------  Details of Selected Knowledge Check -----------------------------------------------
@login_required
def kc_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        ques_list = ""
        ques_option_list = ""
        conquest_id = request.POST.get('id')

        conq_details = []
        conq_list = conquest_list.objects.filter(id = conquest_id)

        total_r = conquest_recipients.objects.filter(conquest_id = conquest_id).count()
        ack_r = conquest_recipients.objects.filter(conquest_id = conquest_id).exclude(attempts="0").count()

        if ack_r == 0:
            c_rate = 0.0
        else:
            c_rate = round(float((ack_r / total_r) * 100), 2)

        set = {}
        set['id']               = conq_list[0].id
        set['title']            = conq_list[0].title
        set['timestamp']        = conq_list[0].timestamp
        set['sub_process']      = conq_list[0].sub_process
        set['questions_count']  = conq_list[0].questions_count
        set['passing_score']    = conq_list[0].passing_score
        set['completion_rate']  = c_rate
        set['max_attempts']     = conq_list[0].max_attempts
        set['sender']           = conq_list[0].sender
        set['deadline']         = conq_list[0].deadline
        conq_details.append(set)

        data['conquest_details'] = conq_details

        conquest_user_score = []
        user_score = conquest_recipients.objects.filter(conquest_id = conquest_id, user_email=request.user.username)

        if user_score:
            set = {}
            set['conquest_id']  = user_score[0].conquest_id
            set['user_email']   = user_score[0].user_email
            set['attempts']     = user_score[0].attempts
            set['score']        = user_score[0].score
            set['status']       = user_score[0].status
            conquest_user_score.append(set)

        data['conquest_user_score'] = conquest_user_score

        ques_id_list = list(conquest_questions.objects.filter(conquest_id=conquest_id))
        random.shuffle(ques_id_list)

        for q in ques_id_list:
            ques_info = question_list.objects.filter(id=q.question_id)
            ques_list = list(chain(ques_list, ques_info))

            ques_op = question_details.objects.filter(question_id=q.question_id)
            ques_option_list = list(chain(ques_option_list, ques_op))

        data['question_details'] = []
        for q in ques_list:
            set                 = {}
            set['id']           = q.id
            set['timestamp']    = q.timestamp
            set['question']     = q.question
            set['type']         = q.type
            set['marks']        = q.marks
            set['media_type']   = q.media_type
            set['media_url']    = q.media_url
            data['question_details'].append(set)

        data['question_op_list'] = []
        for op in ques_option_list:
            set = {}
            set['question_id']      = op.question_id
            set['question_option']  = op.question_option
            set['question_answer']  = op.question_answer
            data['question_op_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE ANSWERS OF KNOWLEDGE TEST GIVEN BY EMPLOYEE -----------------------------------
def save_user_kc_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        conquest_id = request.POST.get('conquest_id')
        kc_sender = request.POST.get('sender')
        kc_reciever = request.user.username
        answers = json.loads(request.POST['answers'])
        timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')

        answer_list = []

        subj = ""
        for check in answers:
            if check['question_type'] == "Subjective":
                subj = "yes"

        ques_id_list = conquest_questions.objects.filter(conquest_id=conquest_id)

        for q in ques_id_list:
            ques_op = question_details.objects.filter(question_id=q.question_id)
            count = 0
            set = {}
            answ = []
            for q_op in ques_op:
                count = count + 1
                if q_op.question_answer == "yes":
                    answ.append(count)

            set['question_id'] = q.question_id
            set['question_answer'] = answ
            answer_list.append(set)

        total_marks = 0
        max_marks = 0
        atmpts = conquest_recipients.objects.filter(conquest_id=conquest_id,
                                                    user_email=request.user.username)
        attempts = int(atmpts[0].attempts) + 1

        for i in answers:
            save_user_ack_answers(conquest_id, i['question_id'], kc_reciever, attempts, i['question_answer'])

        if subj != "yes":
            for i in answers:
                if i['question_type'] == "MAQs":
                    for j in answer_list:
                        if i['question_id'] == j['question_id']:
                            q_max_marks = int(i['question_marks'])
                            t_q_op = question_details.objects.filter(question_id=i['question_id']).count()
                            t_q_correct_op = len(j['question_answer'])
                            t_u_op_selected = len(i['question_answer'])
                            t_u_correct_op = 0
                            t_u_incorrect_op = 0
                            if t_u_op_selected > t_q_correct_op:
                                t_extra_op_selected = t_u_op_selected - t_q_correct_op
                            else:
                                t_extra_op_selected = 0

                            if t_q_op != t_u_op_selected:
                                for op_eval in range(t_u_op_selected):
                                    if i['question_answer'][op_eval] in j['question_answer']:
                                        t_u_correct_op = t_u_correct_op + 1
                                    else:
                                        t_u_incorrect_op = t_u_incorrect_op + 1

                                user_q_marks = (((t_u_op_selected - t_u_incorrect_op) * q_max_marks) / t_q_correct_op) \
                                               - (t_extra_op_selected / t_q_op)

                                total_marks = total_marks + user_q_marks

                            max_marks = max_marks + int(i['question_marks'])
                else:
                    for j in answer_list:
                        if i['question_id'] == j['question_id']:
                            if (i['question_answer'] == j['question_answer']):
                                total_marks = total_marks + int(i['question_marks'])

                            max_marks = max_marks + int(i['question_marks'])

            u_score = round((round(float(total_marks), 2) / round(float(max_marks), 2)), 2) * 100

            conquest_recipients.objects \
                .filter(conquest_id = conquest_id, user_email=request.user.username) \
                .update(attempts=attempts, score=u_score, status="checked" , timestamp = timestamp)

        else:
            conquest_recipients.objects \
                .filter(conquest_id=conquest_id, user_email=request.user.username) \
                .update(attempts=attempts, status="pending", timestamp = timestamp)

        send_knowledge_check_acknowledgement_email(kc_reciever, kc_sender)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END My Tasks - Knowledge Check List ------------------------------------------------





# --------------------------------- Knowledge Check Evaluation ---------------------------------------------------------


# --------------------------------- REDIRECT TO Knowledge Check Evaluation ---------------------------------------------
@login_required
def knowledge_check_evaluation(request):
    request.session.set_expiry(3600)
    account_list = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

    sp_list = ""
    for x in account_list:
        sp = sub_process_list.objects.filter(account_name=x['account_name'])
        sp_list = list(chain(sp_list, sp))

    kc_id_list = []
    for x in sp_list:
        kc_id = conquest_list.objects.filter(sub_process=x.sub_process_name).values('id', 'title').order_by('-id')
        for i in kc_id:
            kc_recipients_count = conquest_recipients.objects.filter(conquest_id = i['id'], status="pending").count()
            if kc_recipients_count > 0:
                set = {}
                set['id'] = i['id']
                set['title'] = i['title']
                kc_id_list.append(set)

    data = {}
    data['conquest_list'] = kc_id_list
    return render(request, 'b4-knowledge-check-evaluation.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Users List of Selected Knowledge Check Evaluation ----------------------------------
def conquest_user_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        title = request.POST.get('conquest_id')
        conquest_data = conquest_list.objects.filter(title=title)
        kc_id = conquest_data[0].id

        data['conquest_user_list'] = []
        users_list = conquest_recipients.objects.filter(conquest_id=kc_id)
        for u in users_list:
            set = {}
            set['conquest_id']  = u.conquest_id
            set['timestamp']    = u.timestamp
            set['user_email']   = u.user_email
            set['attempts']     = u.attempts
            set['score']        = round( float(u.score), 2)
            set['status']       = u.status
            data['conquest_user_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Users List of Selected Knowledge Check Evaluation ----------------------------------
def conquest_user_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_email = request.POST.get('user_email')

        title = request.POST.get('conquest_id')
        conquest_data = conquest_list.objects.filter(title=title)

        kc_id = conquest_data[0].id
        kc_details = conquest_list.objects.filter(pk=kc_id)
        kc_user_details = conquest_recipients.objects.filter(conquest_id=kc_id, user_email=user_email)
        attempt_no = kc_user_details[0].attempts
        kc_user_q_details = conquest_recipients_answers.objects.filter(conquest_id=kc_id, user_email=user_email,
                                                                       attempt_no=attempt_no).order_by('question_id')
        ques_id_list = list(conquest_questions.objects.filter(conquest_id=kc_id))

        data['question_details']    = []
        data['question_op_list']    = []
        data['kc_details']          = []
        data['kc_user_q_details']   = []
        data['kc_user_details']     = []

        for kc in kc_details:

            total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
            ack_r = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate = round(float((ack_r / total_r) * 100), 2)

            set = {}
            set['id']               = kc.id
            set['title']            = kc.title
            set['timestamp']        = kc.timestamp
            set['sub_process']      = kc.sub_process
            set['questions_count']  = kc.questions_count
            set['passing_score']    = kc.passing_score
            set['max_attempts']     = kc.max_attempts
            set['sender']           = kc.sender
            set['deadline']         = kc.deadline
            set['completion_rate']  = c_rate
            data['kc_details'].append(set)

        for q in kc_user_q_details:
            set = {}
            set['conquest_id']  = q.conquest_id
            set['question_id']  = q.question_id
            set['user_email']   = q.user_email
            set['attempt_no']   = q.attempt_no
            set['answers']      = q.answers
            data['kc_user_q_details'].append(set)

        for q in kc_user_details:
            set = {}
            set['conquest_id']  = q.conquest_id
            set['user_email']   = q.user_email
            set['attempts']     = q.attempts
            set['score']        = round( float(q.score), 2)
            set['status']       = q.status
            data['kc_user_details'].append(set)

        for q in ques_id_list:

            ques_info = question_list.objects.filter(id=q.question_id)
            set = {}
            set['id']           = q.question_id
            set['timestamp']    = ques_info[0].timestamp
            set['question']     = ques_info[0].question
            set['type']         = ques_info[0].type
            set['marks']        = ques_info[0].marks
            set['media_type']   = ques_info[0].media_type
            set['media_url']    = ques_info[0].media_url
            data['question_details'].append(set)

            ques_op = question_details.objects.filter(question_id=q.question_id)
            for q_op in ques_op:
                set = {}
                set['question_id']      = q_op.question_id
                set['question_option']  = q_op.question_option
                set['question_answer']  = q_op.question_answer
                data['question_op_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNconquest_user_listCTION -----------------------------------------------------------------------


# --------------------------------- Users submitted Knowledge Check Evaluation Record Update ---------------------------
def user_kc_eval(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        title = request.POST.get('conquest_id')
        conquest_data = conquest_list.objects.filter(title=title)
        kc_id = conquest_data[0].id

        agent_email = request.POST.get('agent_id')
        agent_score = request.POST.get('agent_score')
        sender = request.user.username

        conquest_recipients.objects.filter(conquest_id=kc_id, user_email=agent_email).update(score=agent_score,
                                                                                             status="checked")

        send_kc_eval_check_email(sender, agent_email)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Knowledge Check Evaluation -----------------------------------------------------





# --------------------------------- Elixir -----------------------------------------------------------------------------


# --------------------------------- REDIRECT TO Elixir -----------------------------------------------------------------
@login_required
def elixir(request):
    request.session.set_expiry(3600)
    usrname = request.user.username
    user_type = request.user.email

    if user_type == "superuser" or user_type == "leader":
        project_name = projects_list.objects.all().values('project_name').distinct()

    elif user_type == "admin":
        project_name = []
        acc_list = user_details.objects.filter(email_id = usrname).values('account_name').distinct()
        for ac in acc_list:
            project = projects_list.objects.filter(account_name = ac['account_name']).values('project_name').distinct()
            project_name.append(project)
        project_name = project_name[0]

    else:
        project_name = user_details.objects.filter(email_id = usrname).values('project_name').distinct()

    sp_list = []
    for x in project_name:
        sp = sub_process_list.objects.filter(project_name=x['project_name']).values('sub_process_name')
        for y in sp:
            sp_set = {}
            sp_set['sub_process_name'] = y['sub_process_name']
            sp_list.append(sp_set)

    update = []
    poc_list = []
    for x in sp_list:
        up_list = update_details.objects.filter(sub_process=x['sub_process_name'])
        for up in up_list:
            total_r     = update_recipients.objects.filter(update_id = up.id).count()
            ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

            if ack_r == 0:
                c_rate  = 0.0
            else:
                c_rate  = round(float((ack_r / total_r) * 100), 2)

            set = {}
            set['id']               = up.id
            set['title']            = up.title
            set['timestamp']        = up.timestamp
            set['deadline']         = up.deadline
            set['sub_process']      = up.sub_process
            set['update_type']      = up.update_type
            set['update_source']    = up.update_source
            set['update_message']   = up.update_message
            set['sender']           = up.sender
            set['completion_rate']  = c_rate
            update.append(set)

        ldap_names = update_details.objects.filter(sub_process=x['sub_process_name']).values('sender').distinct()
        for ldap in ldap_names:
            poc_list.append(ldap['sender'])

    ldap_list = []
    for x in poc_list:
        if x not in ldap_list:
            ldap_list.append(x)

    ldap = []
    for poc in ldap_list:
        ldap_set = {}
        ldap_set['sender'] = poc
        ldap.append(ldap_set)

    sender_data = []
    for x in ldap:
        set = {}
        total_updates       = 0
        completed_updates   = 0
        comp_rate           = 0.0

        update_list = update_details.objects.filter(sender=x['sender'])
        for up in update_list:
            total_r = update_recipients.objects.filter(update_id=up.id).count()
            ack_r = update_recipients.objects.filter(update_id=up.id).exclude(ack_status="NA").count()

            total_updates   = total_updates + 1

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate          = round(float((ack_r / total_r) * 100), 2)

            comp_rate       = comp_rate + c_rate

            if c_rate == 100.0:
                completed_updates = completed_updates + 1

        pending_updates = total_updates - completed_updates

        if comp_rate == 0:
            comp_rate = 0.0
        else:
            comp_rate               = round((comp_rate/float(total_updates)), 2)

        set['sender']               = x['sender']
        set['pending_updates']      = pending_updates
        set['completed_updates']    = completed_updates
        set['total_updates']        = total_updates
        set['completion_rate']      = comp_rate
        sender_data.append(set)

    data = {}
    data['update_details'] = update
    data['sender_details'] = sender_data

    return render(request, 'c1-elixir.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Fetch Selected Update Details ------------------------------------------------------
@login_required
def update_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        update_id = request.POST.get('id')

        data['update_details']      = []
        data['update_user_details'] = []

        up_details = update_details.objects.filter(pk=update_id)
        for up in up_details:

            total_r = update_recipients.objects.filter(update_id = up.id).count()
            ack_r = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate = round(float((ack_r / total_r) * 100), 2)

            set = {}
            set['id']               = up.id
            set['title']            = up.title
            set['timestamp']        = up.timestamp
            set['deadline']         = up.deadline
            set['sub_process']      = up.sub_process
            set['update_type']      = up.update_type
            set['update_source']    = up.update_source
            set['update_message']   = up.update_message
            set['sender']           = up.sender
            set['completion_rate']  = c_rate
            data['update_details'].append(set)

        usr_details = update_recipients.objects.filter(update_id = update_id)
        for user_up in usr_details:

            if user_details.objects.filter(email_id=user_up.user_email):
                emp_id = user_details.objects.filter(email_id=user_up.user_email)[0].emp_id
            else:
                emp_id = "Removed"

            if User.objects.filter(username=user_up.user_email).values('first_name'):
                emp_f_name = User.objects.filter(username=user_up.user_email).values('first_name')[0]['first_name']
            else:
                emp_f_name = "Removed"

            if User.objects.filter(username=user_up.user_email).values('last_name'):
                emp_l_name = User.objects.filter(username=user_up.user_email).values('last_name')[0]['last_name']
            else:
                emp_l_name = "Removed"

            if emp_f_name == "Removed":
                emp_name = "Removed"
            else:
                emp_name = emp_f_name + " " + emp_l_name

            set = {}
            set['update_id']    = user_up.update_id
            set['timestamp']    = user_up.timestamp
            set['emp_id']       = emp_id
            set['emp_name']     = emp_name
            set['user_email']   = user_up.user_email
            set['ack_status']   = user_up.ack_status
            set['ack_comment']  = user_up.ack_comment
            data['update_user_details'].append(set)

        conquest_id = update_conquest.objects.filter(update_id=update_id)

        if conquest_id:
            data['conquest_details']    = []
            data['conquest_user_score'] = []
            data['question_details']    = []
            data['question_op_list']    = []

            conq_status = "yes"
            conq_details = conquest_list.objects.filter(id = conquest_id[0].conquest_id)
            for kc in conq_details:

                total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                ack_r = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                if ack_r == 0:
                    c_rate = 0.0
                else:
                    c_rate = round(float((ack_r / total_r) * 100), 2)

                set = {}
                set['id']               = kc.id
                set['title']            = kc.title
                set['timestamp']        = kc.timestamp
                set['sub_process']      = kc.sub_process
                set['questions_count']  = kc.questions_count
                set['passing_score']    = kc.passing_score
                set['max_attempts']     = kc.max_attempts
                set['sender']           = kc.sender
                set['deadline']         = kc.deadline
                set['completion_rate']  = c_rate
                data['conquest_details'].append(set)

            user_score = conquest_recipients.objects.filter(conquest_id = conquest_id[0].conquest_id, user_email=request.user.username)
            for kc in user_score:
                set = {}
                set['conquest_id']  = kc.conquest_id
                set['timestamp']    = kc.timestamp
                set['user_email']   = kc.user_email
                set['attempts']     = kc.attempts
                set['score']        = round( float(kc.score), 2)
                set['status']       = kc.status
                data['conquest_user_score'].append(set)

            ques_id_list = list(conquest_questions.objects.filter(conquest_id=conquest_id[0].conquest_id))
            random.shuffle(ques_id_list)

            for q in ques_id_list:
                ques_info = question_list.objects.filter(id=q.question_id)
                for q_details in ques_info:
                    set = {}
                    set['id']           = q_details.id
                    set['timestamp']    = q_details.timestamp
                    set['question']     = q_details.question
                    set['type']         = q_details.type
                    set['marks']        = q_details.marks
                    set['media_type']   = q_details.media_type
                    set['media_url']    = q_details.media_url
                    data['question_details'].append(set)

                ques_op = question_details.objects.filter(question_id=q.question_id)
                for q_op in ques_op:
                    set = {}
                    set['question_id']      = q_op.question_id
                    set['question_option']  = q_op.question_option
                    set['question_answer']  = q_op.question_answer
                    data['question_op_list'].append(set)

        else:
            conq_status = "no"

        data['conquest'] = conq_status

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FETCH LIST OF UPDATES OF Selected POC ----------------------------------------------
@login_required
def update_poc_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_email = request.POST.get('user_email')

        data['update_details'] = []
        updates_list = update_details.objects.filter(sender=user_email)
        for up in updates_list:

            total_r     = update_recipients.objects.filter(update_id = up.id).count()
            ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

            if ack_r == 0:
                c_rate  = 0.0
            else:
                c_rate  = round(float((ack_r / total_r) * 100), 2)

            set = {}
            set['id']               = up.id
            set['timestamp']        = up.timestamp
            set['title']            = up.title
            set['sub_process']      = up.sub_process
            set['update_type']      = up.update_type
            set['update_source']    = up.update_source
            set['update_message']   = up.update_message
            set['sender']           = up.sender
            set['completion_rate']  = c_rate
            data['update_details'].append(set)

        count = 0
        completion_rate = 0.0
        ack_count = 0
        for x in updates_list:
            count       = count + 1

            total_r     = update_recipients.objects.filter(update_id = x.id).count()
            ack_r       = update_recipients.objects.filter(update_id = x.id).exclude(ack_status="NA").count()

            if ack_r == 0:
                comp_rate = 0.0
            else:
                comp_rate   = round(float((ack_r / total_r) * 100), 2)

            if comp_rate == 100.0:
                ack_count = ack_count + 1

            completion_rate = completion_rate + comp_rate

        if completion_rate == 0:
            completion_rate = 0.0
        else:
            completion_rate     = round((completion_rate / float(count)), 2)

        data['poc_name']        = user_email
        data['completion_rate'] = completion_rate
        data['ack_count']       = ack_count
        data['pend_count']      = count - ack_count

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- List of SubCategories for Selected Category ----------------------------------------
@login_required
def elixir_sub_categories(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        category = request.POST.get('category_name')
        user_type = request.user.email

        acc_list        = []
        proj_list       = []
        sp_list         = []
        up_source_list  = []
        up_type_list    = []

        acc = []
        if user_type == "superuser" or user_type == "leader":
            acc = account_list.objects.all().values('account_name')

        elif user_type == "admin" or user_type == "poc":
            acc = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

        else:
            pass

        if acc != []:
            for i in acc:
                acc_list.append(i['account_name'])

                proj = []
                if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                    proj = projects_list.objects.filter(account_name=i['account_name']).values('project_name')

                elif user_type == "poc":
                    proj = user_details.objects.filter(email_id=request.user.username, account_name=i['account_name']) \
                        .values('project_name').distinct()

                else:
                    pass

                if proj != []:
                    for j in proj:
                        proj_list.append(j['project_name'])

                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')

                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                sp_list.append(k['sub_process_name'])

        source_list = update_source_list.objects.all().values('update_source_name')
        for s in source_list:
            up_source_list.append(s['update_source_name'])

        type_list = update_type_list.objects.all().values('update_type_name')
        for t in type_list:
            up_type_list.append(t['update_type_name'])

        if category == "Account":
            data['sub_category'] = acc_list

        if category == "Project":
            data['sub_category'] = proj_list

        if category == "Sub Process":
            data['sub_category'] = sp_list

        if category == "Update Source":
            data['sub_category'] = up_source_list

        if category == "Update Type":
            data['sub_category'] = up_type_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Elixir List with Filtered Data -----------------------------------------------------
@login_required
def elixir_filtered_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_type = request.user.email

        category = request.POST.get('category_name')
        subcategory = request.POST.get('subcategory_name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        update = ""
        updates_list = []

        if start_date == "na":

            if category == "Account":
                proj = []
                if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                    proj = projects_list.objects.filter(account_name = subcategory).values('project_name')
                elif user_type == "poc":
                    proj = user_details.objects.filter(email_id=request.user.username, account_name = subcategory) \
                        .values('project_name').distinct()
                else:
                    pass

                if proj != []:
                    for j in proj:
                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')
                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                up_list = update_details.objects.filter(sub_process = k['sub_process_name'])
                                update = list(chain(update, up_list))

            if category == "Project":
                sp = []
                if user_type != "agent":
                    sp = sub_process_list.objects.filter(project_name = subcategory).values('sub_process_name')
                else:
                    pass

                if sp != []:
                    for k in sp:
                        up_list = update_details.objects.filter(sub_process=k['sub_process_name'])
                        update = list(chain(update, up_list))

            if category == "Sub Process":
                up_list = update_details.objects.filter(sub_process = subcategory)
                update = list(chain(update, up_list))

            if category == "Update Source":
                acc = []
                if user_type == "superuser" or user_type == "leader":
                    acc = account_list.objects.all().values('account_name')
                elif user_type == "admin" or user_type == "poc":
                    acc = user_details.objects.filter(email_id=request.user.username).values(
                        'account_name').distinct()
                else:
                    pass

                if acc != []:
                    for i in acc:
                        proj = []
                        if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                            proj = projects_list.objects.filter(account_name=i['account_name']).values('project_name')
                        elif user_type == "poc":
                            proj = user_details.objects.filter(email_id=request.user.username,
                                                               account_name=i['account_name']) \
                                .values('project_name').distinct()
                        else:
                            pass

                        if proj != []:
                            for j in proj:
                                sp = []
                                if user_type != "agent":
                                    sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                        'sub_process_name')
                                else:
                                    pass

                                if sp != []:
                                    for k in sp:
                                        up_list = update_details.objects.filter(sub_process = k['sub_process_name'], update_source = subcategory)
                                        update = list(chain(update, up_list))

            if category == "Update Type":
                acc = []
                if user_type == "superuser" or user_type == "leader":
                    acc = account_list.objects.all().values('account_name')
                elif user_type == "admin" or user_type == "poc":
                    acc = user_details.objects.filter(email_id=request.user.username).values(
                        'account_name').distinct()
                else:
                    pass

                if acc != []:
                    for i in acc:
                        proj = []
                        if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                            proj = projects_list.objects.filter(account_name=i['account_name']).values('project_name')
                        elif user_type == "poc":
                            proj = user_details.objects.filter(email_id=request.user.username,
                                                               account_name=i['account_name']) \
                                .values('project_name').distinct()
                        else:
                            pass

                        if proj != []:
                            for j in proj:
                                sp = []
                                if user_type != "agent":
                                    sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                        'sub_process_name')
                                else:
                                    pass

                                if sp != []:
                                    for k in sp:
                                        up_list = update_details.objects.filter(sub_process = k['sub_process_name'], update_type = subcategory)
                                        update = list(chain(update, up_list))

            for up in update:

                total_r     = update_recipients.objects.filter(update_id = up.id).count()
                ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                if ack_r == 0:
                    comp_rate = 0.0
                else:
                    comp_rate   = round(float((ack_r / total_r) * 100), 2)

                set = {}
                set['id']               = up.id
                set['timestamp']        = up.timestamp
                set['title']            = up.title
                set['update_message']   = up.update_message
                set['sub_process']      = up.sub_process
                set['update_source']    = up.update_source
                set['update_type']      = up.update_type
                set['sender']           = up.sender
                set['deadline']         = up.deadline
                set['completion_rate']  = comp_rate
                updates_list.append(set)

        else:

            if category == "Account":
                proj = []
                if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                    proj = projects_list.objects.filter(account_name=subcategory).values('project_name')
                elif user_type == "poc":
                    proj = user_details.objects.filter(email_id=request.user.username, account_name=subcategory) \
                        .values('project_name').distinct()
                else:
                    pass

                if proj != []:
                    for j in proj:
                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')
                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                up_list = update_details.objects.filter(sub_process=k['sub_process_name'])
                                for up in up_list:
                                    val1 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                                        start_date, "%d-%m-%Y")
                                    val2 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                      "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                                "%d-%m-%Y")
                                    if val1 and val2:

                                        total_r     = update_recipients.objects.filter(update_id = up.id).count()
                                        ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                                        if ack_r == 0:
                                            comp_rate = 0.0
                                        else:
                                            comp_rate   = round(float((ack_r / total_r) * 100), 2)

                                        set = {}
                                        set['id']               = up.id
                                        set['timestamp']        = up.timestamp
                                        set['title']            = up.title
                                        set['update_message']   = up.update_message
                                        set['sub_process']      = up.sub_process
                                        set['update_source']    = up.update_source
                                        set['update_type']      = up.update_type
                                        set['sender']           = up.sender
                                        set['deadline']         = up.deadline
                                        set['completion_rate']  = comp_rate
                                        updates_list.append(set)

            if category == "Project":
                sp = []
                if user_type != "agent":
                    sp = sub_process_list.objects.filter(project_name=subcategory).values('sub_process_name')
                else:
                    pass

                if sp != []:
                    for k in sp:
                        up_list = update_details.objects.filter(sub_process=k['sub_process_name'])
                        for up in up_list:
                            val1 = datetime.datetime.strptime(up.timestamp[0:10],
                                                              "%d-%m-%Y") >= datetime.datetime.strptime(
                                start_date, "%d-%m-%Y")
                            val2 = datetime.datetime.strptime(up.timestamp[0:10],
                                                              "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                        "%d-%m-%Y")
                            if val1 and val2:

                                total_r     = update_recipients.objects.filter(update_id = up.id).count()
                                ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                                if ack_r == 0:
                                    comp_rate = 0.0
                                else:
                                    comp_rate   = round(float((ack_r / total_r) * 100), 2)

                                set = {}
                                set['id']               = up.id
                                set['timestamp']        = up.timestamp
                                set['title']            = up.title
                                set['update_message']   = up.update_message
                                set['sub_process']      = up.sub_process
                                set['update_source']    = up.update_source
                                set['update_type']      = up.update_type
                                set['sender']           = up.sender
                                set['deadline']         = up.deadline
                                set['completion_rate']  = comp_rate
                                updates_list.append(set)

            if category == "Sub Process":
                up_list = update_details.objects.filter(sub_process=subcategory)
                for up in up_list:
                    val1 = datetime.datetime.strptime(up.timestamp[0:10],
                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date, "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(up.timestamp[0:10],
                                                      "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                "%d-%m-%Y")
                    if val1 and val2:

                        total_r     = update_recipients.objects.filter(update_id = up.id).count()
                        ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                        if ack_r == 0:
                            comp_rate = 0.0
                        else:
                            comp_rate   = round(float((ack_r / total_r) * 100), 2)

                        set = {}
                        set['id']               = up.id
                        set['timestamp']        = up.timestamp
                        set['title']            = up.title
                        set['update_message']   = up.update_message
                        set['sub_process']      = up.sub_process
                        set['update_source']    = up.update_source
                        set['update_type']      = up.update_type
                        set['sender']           = up.sender
                        set['deadline']         = up.deadline
                        set['completion_rate']  = comp_rate
                        updates_list.append(set)

            if category == "Update Source":
                acc = []
                if user_type == "superuser" or user_type == "leader":
                    acc = account_list.objects.all().values('account_name')
                elif user_type == "admin" or user_type == "poc":
                    acc = user_details.objects.filter(email_id=request.user.username).values(
                        'account_name').distinct()
                else:
                    pass

                if acc != []:
                    for i in acc:
                        proj = []
                        if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                            proj = projects_list.objects.filter(account_name=i['account_name']).values('project_name')
                        elif user_type == "poc":
                            proj = user_details.objects.filter(email_id=request.user.username,
                                                               account_name=i['account_name']) \
                                .values('project_name').distinct()
                        else:
                            pass

                        if proj != []:
                            for j in proj:
                                sp = []
                                if user_type != "agent":
                                    sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                        'sub_process_name')
                                else:
                                    pass

                                if sp != []:
                                    for k in sp:
                                        up_list = update_details.objects.filter(sub_process=k['sub_process_name'],
                                                                                update_source=subcategory)
                                        for up in up_list:
                                            val1 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                              "%d-%m-%Y") >= datetime.datetime.strptime(
                                                start_date, "%d-%m-%Y")
                                            val2 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                              "%d-%m-%Y") <= datetime.datetime.strptime(
                                                end_date,
                                                "%d-%m-%Y")
                                            if val1 and val2:

                                                total_r     = update_recipients.objects.filter(update_id = up.id).count()
                                                ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                                                if ack_r == 0:
                                                    comp_rate = 0.0
                                                else:
                                                    comp_rate   = round(float((ack_r / total_r) * 100), 2)

                                                set = {}
                                                set['id']               = up.id
                                                set['timestamp']        = up.timestamp
                                                set['title']            = up.title
                                                set['update_message']   = up.update_message
                                                set['sub_process']      = up.sub_process
                                                set['update_source']    = up.update_source
                                                set['update_type']      = up.update_type
                                                set['sender']           = up.sender
                                                set['deadline']         = up.deadline
                                                set['completion_rate']  = comp_rate
                                                updates_list.append(set)

            if category == "Update Type":
                acc = []
                if user_type == "superuser" or user_type == "leader":
                    acc = account_list.objects.all().values('account_name')
                elif user_type == "admin" or user_type == "poc":
                    acc = user_details.objects.filter(email_id=request.user.username).values(
                        'account_name').distinct()
                else:
                    pass

                if acc != []:
                    for i in acc:
                        proj = []
                        if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                            proj = projects_list.objects.filter(account_name=i['account_name']).values('project_name')
                        elif user_type == "poc":
                            proj = user_details.objects.filter(email_id=request.user.username,
                                                               account_name=i['account_name']) \
                                .values('project_name').distinct()
                        else:
                            pass

                        if proj != []:
                            for j in proj:
                                sp = []
                                if user_type != "agent":
                                    sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                        'sub_process_name')
                                else:
                                    pass

                                if sp != []:
                                    for k in sp:
                                        up_list = update_details.objects.filter(sub_process=k['sub_process_name'],
                                                                                update_type=subcategory)
                                        for up in up_list:
                                            val1 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                              "%d-%m-%Y") >= datetime.datetime.strptime(
                                                start_date, "%d-%m-%Y")
                                            val2 = datetime.datetime.strptime(up.timestamp[0:10],
                                                                              "%d-%m-%Y") <= datetime.datetime.strptime(
                                                end_date,
                                                "%d-%m-%Y")
                                            if val1 and val2:

                                                total_r     = update_recipients.objects.filter(update_id = up.id).count()
                                                ack_r       = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                                                if ack_r == 0:
                                                    comp_rate = 0.0
                                                else:
                                                    comp_rate   = round(float((ack_r / total_r) * 100), 2)

                                                set = {}
                                                set['id']               = up.id
                                                set['timestamp']        = up.timestamp
                                                set['title']            = up.title
                                                set['update_message']   = up.update_message
                                                set['sub_process']      = up.sub_process
                                                set['update_source']    = up.update_source
                                                set['update_type']      = up.update_type
                                                set['sender']           = up.sender
                                                set['deadline']         = up.deadline
                                                set['completion_rate']  = comp_rate
                                                updates_list.append(set)

        data['update_details'] = updates_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Change Update validity -------------------------------------------------------------
@login_required
def change_update_validity(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        updates_list = json.loads(request.POST['updates_list'])
        validity = request.POST.get('validity')

        updates_count = len(updates_list)

        for i in range(updates_count):
            update_details.objects.filter(pk = updates_list[i]).update(deadline = validity)


        data['message'] = "Validity Changed Successfully"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Elixir -------------------------------------------------------------------------





# --------------------------------- Create New Update ------------------------------------------------------------------


# --------------------------------- REDIRECT TO CREATE UPDATE PAGE -----------------------------------------------------
@login_required
def createUpdate(request):
    request.session.set_expiry(3600)
    proj_list = user_details.objects.filter(email_id=request.user.username).values('project_name').distinct()

    sub_process = ""
    for x in proj_list:
        grp = sub_process_list.objects.filter(project_name=x['project_name'])
        sub_process = list(chain(sub_process, grp))

    update_type = update_type_list.objects.all()
    update_source = update_source_list.objects.all()

    data = {}
    data['sub_process'] = sub_process
    data['update_editor'] = update_editorForm()
    data['update_type'] = update_type
    data['update_source'] = update_source

    return render(request, 'c2-create-new-update.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- UPLOAD ATTACHMENT ------------------------------------------------------------------
def upload_attachment_file(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        acc_name = user_details.objects.filter(email_id=request.user.username).values('account_name')
        myfile = request.FILES.getlist('attachment_file')

        valid_files = 0

        for i in range(len(myfile)):

            file_ext = myfile[i].name.split(".")[-1]

            if  file_ext.lower() == "xls" or \
                file_ext.lower() == "xlsx" or \
                file_ext.lower() == "txt" or \
                file_ext.lower() == "docx" or \
                file_ext.lower() == "doc" or \
                file_ext.lower() == "pdf" :

                valid_files = valid_files + 1


        if len(myfile) == valid_files :

            uploaded_file_url = []

            for c in range(len(myfile)):

                fs = FileSystemStorage('media/attachments/' + acc_name[0]['account_name'].replace(" ", "_") + '/')
                filename = fs.save((myfile[c].name).replace(" ", "_"), myfile[c])
                file_url = fs.url("/attachments/" + acc_name[0]['account_name'].replace(" ", "_") + "/" + filename)
                uploaded_file_url.append(file_url)

            data['code'] = 1
            data['message'] = "Attachment uploaded successfully"
            data['uploaded_media_url'] = uploaded_file_url

        else:
            data['code'] = 0
            data['message'] = "Upload Unsuccessful : Invalid File Type, Please select .txt, excel, word, pdf files only."

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SEND UPDATE ------------------------------------------------------------------------
@login_required
def send_update(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        subject         = request.POST['title']
        sub_process     = json.loads(request.POST['sub_process'])
        update_type     = request.POST['update_type']
        update_source   = request.POST['update_source']
        message         = request.POST['message']
        attachments     = json.loads(request.POST['attachments'])
        sender          = request.user.username
        timestamp       = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        validity        = request.POST['validity']

        kc_score        = request.POST['kc_score']
        kc_attempts     = request.POST['kc_attempts']
        kc_title        = request.POST['kc_title']
        kc_questions    = json.loads(request.POST['kc_questions'])

        if kc_score != "na":
            q_count     = len(kc_questions)
            kc_ques_arr = save_kc_question(kc_questions, timestamp, sender)

        else:
            q_count     = 0
            kc_ques_arr = []

        sub_process_count = len(sub_process)
        for i in range(sub_process_count):

            title = update_source + " - " + update_type + " - " + subject

            update = update_details(title=title,
                                    sub_process=sub_process[i],
                                    update_type=update_type,
                                    update_source=update_source,
                                    update_message=message,
                                    sender=sender,
                                    timestamp=timestamp,
                                    deadline=validity)
            update.save()

            receiver = []
            users_list = sub_process_details.objects.values('user_email').filter(sub_process_name=sub_process[i])
            for usr in users_list:
                up_recipients = update_recipients(update_id=update.id,
                                                  user_email=usr['user_email'],
                                                  ack_status="NA",
                                                  ack_comment="NA",
                                                  timestamp = "NA")
                up_recipients.save()

                receiver.append(usr['user_email'])

            cc_receiver = cc_list_of_poc_and_admins(sub_process[i], sender)
            
            print("sender - ", sender)
            print("Receiver - ", receiver)
            print("cc_receiver - ", cc_receiver)
            if kc_score == "na":
                send_update_mail(sender, receiver, cc_receiver, title, update_source, update_type, message, validity, attachments)
            else:
                conquest_id, agents = save_kc(timestamp, sub_process[i], validity, kc_score, kc_attempts, kc_title, kc_ques_arr, q_count, sender)
                save_up_conq(update.id, conquest_id)
                send_update_kc_mail(sender, receiver, cc_receiver, title, update_source, update_type, message, validity, attachments)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE KNOWLEDGE TEST ----------------------------------------------------------------
@login_required
def send_kc(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        up_id       = request.POST.get('up_id')
        timestamp   = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        sub_process = json.loads(request.POST['sub_process'])
        validity    = request.POST.get('validity')
        score       = request.POST.get('score')
        attempts    = request.POST.get('attempts')
        kc_title    = request.POST.get('kc_title')
        questions   = json.loads(request.POST['questions'])
        q_count     = len(questions)
        sender      = request.user.username

        kc_ques_arr = save_kc_question(questions, timestamp, sender)

        if up_id != "na":
            up_details = update_details.objects.filter(id=up_id)
            subprocess = up_details[0].sub_process
            validity = up_details[0].deadline

            pkt_id, receiver = save_kc(timestamp, subprocess, validity, score, attempts, kc_title, kc_ques_arr, q_count,
                                       sender)

            save_up_conq(up_id, pkt_id)

            cc_receiver = cc_list_of_poc_and_admins(sub_process, sender)
            
            data['message'] = "Knowledge Test Sent successfully for a previously sent Update"
            send_update_pkt_email(sender, receiver, cc_receiver)

        else:
            sub_process_count = len(sub_process)
            for i in range(sub_process_count):

                pkt_id, receiver = save_kc(timestamp, sub_process[i], validity, score, attempts, kc_title, kc_ques_arr, q_count, sender)

                cc_receiver = cc_list_of_poc_and_admins(sub_process[i], sender)

                send_pkt_email(sender, receiver, cc_receiver)

            data['message'] = "Knowledge Test Sent successfully"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)

def save_kc_question(questions, timestamp, sender):
    ques_id_arr = []
    for a in questions:
        question    = a['question']
        type        = a['question_type']
        media_type  = a['question_media_type']
        media_url   = a['question_media_url']
        marks       = a['question_marks']

        q_id = save_question_fn(question, type, timestamp, media_type, media_url, marks, sender)
        ques_id_arr.append(q_id)

        op_c = 0
        for op in a['question_options']:
            ans_c = 0
            for ans in a['question_answer']:
                if op_c == ans_c:
                    question_option = op
                    q_ans_val = ans
                    if a['question_type'] != "Subjective":
                        if q_ans_val == 0:
                            question_answer = "no"
                        else:
                            question_answer = "yes"
                    else:
                        question_answer = q_ans_val
                    save_option_fn(q_id, question_option, question_answer)
                ans_c = ans_c + 1
            op_c = op_c + 1

    return ques_id_arr

def save_kc(timestamp, sub_process, validity, score, attempts, kc_title, kc_ques_arr, q_count, sender):

    pkt_id = save_kc_details(timestamp, sub_process, q_count, score, sender, validity, attempts, kc_title)

    for a in kc_ques_arr:

        conq_ques = conquest_questions(conquest_id=pkt_id,
                                       question_id=a)
        conq_ques.save()

    sbp_users_list = sub_process_details.objects.filter(sub_process_name=sub_process)

    reciever = []
    for usr in sbp_users_list:
        conq_id     = pkt_id
        agent       = usr.user_email
        q_attempts  = "0"
        q_score     = "0"
        q_status    = "na"
        t_stamp     = "NA"
        save_conquest_recipients(conq_id, agent, q_attempts, q_score, q_status, t_stamp)
        reciever.append(agent)

    return pkt_id, reciever

def save_kc_details(timestamp, sub_process, q_count, score, sender, validity, attempts, kc_title):
    pkt = conquest_list(timestamp=timestamp,
                        sub_process=sub_process,
                        questions_count=q_count,
                        passing_score=score,
                        sender=sender,
                        deadline=validity,
                        max_attempts=attempts)
    pkt.save()

    title = "KC"+str(pkt.id)+"-"+kc_title
    conquest_list.objects.filter(id = pkt.id).update(title=title)

    return pkt.id

def save_question_fn(question, type, timestamp, media_type, media_url, marks, creator):
    ques = question_list(question=question,
                         type=type,
                         creator=creator,
                         timestamp=timestamp,
                         media_type=media_type,
                         media_url=media_url,
                         marks=marks)
    ques.save()
    return ques.id

def save_option_fn(question_id, question_option, question_answer):
    option = question_details(question_id=question_id,
                              question_option=question_option,
                              question_answer=question_answer)
    option.save()

def save_conquest_recipients(conq_id, reciever, attempts, score, status, timestamp):
    conq_rec = conquest_recipients(conquest_id=conq_id,
                                   user_email=reciever,
                                   attempts=attempts,
                                   score=score,
                                   status=status,
                                   timestamp = timestamp)
    conq_rec.save()
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE KNOWLEDGE TEST WITH UPDATE ----------------------------------------------------
def save_up_conq(update_id, conquest_id):
    up_conq = update_conquest(update_id=update_id,
                              conquest_id=conquest_id)
    up_conq.save()

    return JsonResponse("Saved", safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Create New Update --------------------------------------------------------------





# --------------------------------- All Updates ------------------------------------------------------------------------


# --------------------------------- REDIRECT TO All UPDATE -------------------------------------------------------------
@login_required
def allUpdate(request):
    request.session.set_expiry(3600)
    account_list = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

    sp_list = ""
    for x in account_list:
        sp = sub_process_list.objects.filter(account_name=x['account_name'])
        sp_list = list(chain(sp_list, sp))

    update_id_list = []
    for x in sp_list:
        update_id = update_details.objects.filter(sub_process=x.sub_process_name).order_by('-id')
        for id in update_id:
            update_id_list.append(id.id)

    update_id_list.sort(reverse=True)

    updates_list = ""
    acknw_status = ""
    for id in update_id_list:
        upd = update_details.objects.filter(pk=id)
        updates_list = list(chain(updates_list, upd))
        ack = update_recipients.objects.filter(user_email=request.user.username, update_id=id)
        acknw_status = list(chain(acknw_status, ack))

    data = {}
    data['updates_details'] = updates_list
    data['ack_status'] = acknw_status
    return render(request, 'c3-all-update.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER UPDATES AS PER DATE RANGE SELECTED ------------------------------------------
def allupdates_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        account_list = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

        sp_list = ""
        for x in account_list:
            sp = sub_process_list.objects.filter(account_name=x['account_name'])
            sp_list = list(chain(sp_list, sp))

        update_id_list = []
        for x in sp_list:
            update_id = update_details.objects.filter(sub_process=x.sub_process_name).order_by('-id')
            for id in update_id:
                update_id_list.append(id.id)

        update_id_list.sort(reverse=True)

        data['update_details']  = []
        data['ack_status']      = []
        for id in update_id_list:
            upd = update_details.objects.filter(pk=id)
            for y in upd:
                val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                    start_date, "%d-%m-%Y")
                val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                               "%d-%m-%Y")
                if val1:
                    if val2:

                        for up in upd:

                            total_r         = update_recipients.objects.filter(update_id = up.id).count()
                            ack_r           = update_recipients.objects.filter(update_id = up.id).exclude(ack_status="NA").count()

                            if ack_r == 0:
                                comp_rate   = 0.0
                            else:
                                comp_rate   = round(float((ack_r / total_r) * 100), 2)

                            set = {}
                            set['id']               = up.id
                            set['timestamp']        = up.timestamp
                            set['title']            = up.title
                            set['deadline']         = up.deadline
                            set['sub_process']      = up.sub_process
                            set['update_type']      = up.update_type
                            set['update_source']    = up.update_source
                            set['update_message']   = up.update_message
                            set['sender']           = up.sender
                            set['completion_rate']  = comp_rate
                            data['update_details'].append(set)

                        up_ack = update_recipients.objects.filter(update_id = id)[:1]
                        for ack in up_ack:
                            set = {}
                            set['update_id']        = ack.update_id
                            set['user_email']       = ack.user_email
                            set['ack_status']       = ack.ack_status
                            set['ack_comment']      = ack.ack_comment
                            data['ack_status'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END All Updates --------------------------------------------------------------------





# --------------------------------- Sent Updates -----------------------------------------------------------------------


# --------------------------------- REDIRECT TO Sent UPDATE ------------------------------------------------------------
@login_required
def sentUpdate(request):
    request.session.set_expiry(3600)
    update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')
    updates_list = ""
    acknw_status = ""
    for x in update_id_list:
        upd = update_details.objects.filter(pk=x.id)
        updates_list = list(chain(updates_list, upd))
        ack = update_recipients.objects.filter(user_email=request.user.username, update_id=x.id)
        acknw_status = list(chain(acknw_status, ack))

    data = {}
    data['updates_details'] = updates_list
    data['ack_status'] = acknw_status
    return render(request, 'c4-sent-update.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER UPDATES AS PER DATE RANGE SELECTED ------------------------------------------
def sentupdates_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')

        data['update_details']  = []
        data['ack_status']      = []
        for x in update_id_list:
            val1 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                           "%d-%m-%Y")
            if val1:
                if val2:

                    total_r = update_recipients.objects.filter(update_id = x.id).count()
                    ack_r = update_recipients.objects.filter(update_id = x.id).exclude(ack_status="NA").count()

                    if ack_r == 0:
                        comp_rate = 0.0
                    else:
                        comp_rate = round(float((ack_r / total_r) * 100), 2)

                    set = {}
                    set['id']               = x.id
                    set['timestamp']        = x.timestamp
                    set['title']            = x.title
                    set['deadline']         = x.deadline
                    set['sub_process']      = x.sub_process
                    set['update_type']      = x.update_type
                    set['update_source']    = x.update_source
                    set['update_message']   = x.update_message
                    set['sender']           = x.sender
                    set['completion_rate']  = comp_rate
                    data['update_details'].append(set)

                    up_ack = update_recipients.objects.filter(update_id=id)[:1]
                    for ack in up_ack:
                        set = {}
                        set['update_id'] = ack.update_id
                        set['user_email'] = ack.user_email
                        set['ack_status'] = ack.ack_status
                        set['ack_comment'] = ack.ack_comment
                        data['ack_status'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Sent Updates -------------------------------------------------------------------





# --------------------------------- Pending Updates --------------------------------------------------------------------


# --------------------------------- REDIRECT TO Pending UPDATE ---------------------------------------------------------
@login_required
def pendingUpdate(request):
    request.session.set_expiry(3600)
    update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')
    updates_list = ""
    acknw_status = ""
    for x in update_id_list:

        total_r     = update_recipients.objects.filter(update_id = x.id).count()
        ack_r       = update_recipients.objects.filter(update_id = x.id).exclude(ack_status="NA").count()

        if ack_r == 0:
            comp_rate = 0.0
        else:
            comp_rate   = round(float((ack_r / total_r) * 100), 2)

        if comp_rate < 100.0:
            upd = update_details.objects.filter(pk = x.id)
            updates_list = list(chain(updates_list, upd))
            ack = update_recipients.objects.filter(user_email=request.user.username, update_id=x.id)
            acknw_status = list(chain(acknw_status, ack))

    data = {}
    data['updates_details'] = updates_list
    data['ack_status'] = acknw_status
    return render(request, 'c5-pending-update.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER UPDATES AS PER DATE RANGE SELECTED ------------------------------------------
def pendingupdates_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')

        data['update_details'] = []
        data['ack_status'] = []
        for x in update_id_list:
            val1 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                           "%d-%m-%Y")
            if val1:
                if val2:

                    total_r = update_recipients.objects.filter(update_id=x.id).count()
                    ack_r = update_recipients.objects.filter(update_id=x.id).exclude(ack_status="NA").count()

                    if ack_r == 0:
                        comp_rate = 0.0
                    else:
                        comp_rate = round(float((ack_r / total_r) * 100), 2)

                    if comp_rate < 100.0:
                        set = {}
                        set['id'] = x.id
                        set['timestamp'] = x.timestamp
                        set['title'] = x.title
                        set['deadline'] = x.deadline
                        set['sub_process'] = x.sub_process
                        set['update_type'] = x.update_type
                        set['update_source'] = x.update_source
                        set['update_message'] = x.update_message
                        set['sender'] = x.sender
                        set['completion_rate'] = comp_rate
                        data['update_details'].append(set)

                        up_ack = update_recipients.objects.filter(update_id=id)[:1]
                        for ack in up_ack:
                            set = {}
                            set['update_id'] = ack.update_id
                            set['user_email'] = ack.user_email
                            set['ack_status'] = ack.ack_status
                            set['ack_comment'] = ack.ack_comment
                            data['ack_status'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Pending Updates ----------------------------------------------------------------





# --------------------------------- Acknowledged Updates ---------------------------------------------------------------


# --------------------------------- REDIRECT TO Acknowledged UPDATE ----------------------------------------------------
@login_required
def ackUpdate(request):
    request.session.set_expiry(3600)
    update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')

    updates_list = ""
    acknw_status = ""
    for x in update_id_list:

        total_r     = update_recipients.objects.filter(update_id = x.id).count()
        ack_r       = update_recipients.objects.filter(update_id = x.id).exclude(ack_status="NA").count()

        if ack_r == 0:
            comp_rate = 0.0
        else:
            comp_rate   = round(float((ack_r / total_r) * 100), 2)

        if comp_rate == 100.0:
            upd = update_details.objects.filter(pk=x.id)
            updates_list = list(chain(updates_list, upd))
            ack = update_recipients.objects.filter(user_email=request.user.username, update_id=x.id)
            acknw_status = list(chain(acknw_status, ack))

    data = {}
    data['updates_details'] = updates_list
    data['ack_status'] = acknw_status
    return render(request, 'c6-acknowledged-update.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER UPDATES AS PER DATE RANGE SELECTED ------------------------------------------
def ackupdates_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        update_id_list = update_details.objects.filter(sender=request.user.username).order_by('-id')

        data['update_details'] = []
        data['ack_status'] = []
        for x in update_id_list:
            val1 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                start_date, "%d-%m-%Y")
            val2 = datetime.datetime.strptime(x.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                           "%d-%m-%Y")
            if val1:
                if val2:

                    total_r = update_recipients.objects.filter(update_id=x.id).count()
                    ack_r = update_recipients.objects.filter(update_id=x.id).exclude(ack_status="NA").count()

                    if ack_r == 0:
                        comp_rate = 0.0
                    else:
                        comp_rate = round(float((ack_r / total_r) * 100), 2)

                    if comp_rate == 100.0:
                        set = {}
                        set['id'] = x.id
                        set['timestamp'] = x.timestamp
                        set['title'] = x.title
                        set['deadline'] = x.deadline
                        set['sub_process'] = x.sub_process
                        set['update_type'] = x.update_type
                        set['update_source'] = x.update_source
                        set['update_message'] = x.update_message
                        set['sender'] = x.sender
                        set['completion_rate'] = comp_rate
                        data['update_details'].append(set)

                        up_ack = update_recipients.objects.filter(update_id=id)[:1]
                        for ack in up_ack:
                            set = {}
                            set['update_id']    = ack.update_id
                            set['user_email']   = ack.user_email
                            set['timestamp']    = ack.timestamp
                            set['ack_status']   = ack.ack_status
                            set['ack_comment']  = ack.ack_comment
                            data['ack_status'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Acknowledged Updates -----------------------------------------------------------





# --------------------------------- Conquest ---------------------------------------------------------------------------


# --------------------------------- REDIRECT TO CONQUEST ---------------------------------------------------------------
@login_required
def conquest(request):
    request.session.set_expiry(3600)
    usrname = request.user.username
    user_type = request.user.email

    if user_type == "superuser" or user_type == "leader":
        project_name = projects_list.objects.all().values('project_name').distinct()

    elif user_type == "admin":
        project_name = []
        acc_list = user_details.objects.filter(email_id = usrname).values('account_name').distinct()
        for ac in acc_list:
            project = projects_list.objects.filter(account_name = ac['account_name']).values('project_name').distinct()
            project_name.append(project)
        project_name = project_name[0]

    else:
        project_name = user_details.objects.filter(email_id = usrname).values('project_name').distinct()

    sp_list = []
    for x in project_name:
        sp = sub_process_list.objects.filter(project_name=x['project_name']).values('sub_process_name')
        for y in sp:
            sp_set = {}
            sp_set['sub_process_name'] = y['sub_process_name']
            sp_list.append(sp_set)

    conquest = []
    for x in sp_list:
        kc_list = conquest_list.objects.filter(sub_process=x['sub_process_name'])

        for y in kc_list:
            val = sub_process_list.objects.filter(sub_process_name=x['sub_process_name'])

            total_r = conquest_recipients.objects.filter(conquest_id = y.id).count()
            ack_r   = conquest_recipients.objects.filter(conquest_id = y.id).exclude(attempts="0").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate  = round(float((ack_r / total_r) * 100), 2)

            object = {}
            object['id']                = y.id
            object['title']             = y.title
            object['sender']            = y.sender
            object['timestamp']         = y.timestamp
            object['sub_process']       = y.sub_process
            object['questions_count']   = y.questions_count
            object['passing_score']     = y.passing_score
            object['max_attempts']      = y.max_attempts
            object['deadline']          = y.deadline
            object['completion_rate']   = c_rate
            object['account']           = val[0].account_name
            object['project']           = val[0].project_name

            conquest.append(object)

    data = {}
    data['conquest'] = conquest
    return render(request, 'd1-conquest.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DETAILS OF SELECTED KNOWLEDGE CHECK ------------------------------------------------
def conquest_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        conquest_id     = request.POST.get('conquest_id')

        conquest_data   = conquest_list.objects.filter(id=conquest_id)
        users_list      = conquest_recipients.objects.filter(conquest_id=conquest_id)

        total_r         = conquest_recipients.objects.filter(conquest_id = conquest_id).count()
        ack_r           = conquest_recipients.objects.filter(conquest_id = conquest_id).exclude(attempts="0").count()

        if ack_r == 0:
            c_rate = 0.0
        else:
            c_rate          = round(float((ack_r / total_r) * 100), 2)

        pending_r       = total_r - ack_r
        total_pass = 0
        total_fail = 0

        data['users_list'] = []

        for u in users_list:
            if int(u.attempts) > 0:
                if float(u.score) < float(conquest_data[0].passing_score):
                    total_fail  = total_fail + 1
                else:
                    total_pass  = total_pass + 1

            if user_details.objects.filter(email_id=u.user_email):
                emp_id      = user_details.objects.filter(email_id=u.user_email)[0].emp_id
            else:
                emp_id = "Removed"

            if User.objects.filter(username=u.user_email).values('first_name'):
                emp_f_name  = User.objects.filter(username=u.user_email).values('first_name')[0]['first_name']
            else:
                emp_f_name = "Removed"

            if User.objects.filter(username=u.user_email).values('last_name'):
                emp_l_name = User.objects.filter(username=u.user_email).values('last_name')[0]['last_name']
            else:
                emp_l_name = "Removed"

            if emp_f_name == "Removed":
                emp_name = "Removed"
            else:
                emp_name = emp_f_name + " " + emp_l_name

            set = {}
            set['conquest_id']  = u.conquest_id
            set['emp_id']       = emp_id
            set['emp_name']     = emp_name
            set['user_email']   = u.user_email
            set['timestamp']   = u.timestamp
            set['attempts']     = u.attempts
            set['score']        = round( float(u.score), 2)
            set['status']       = u.status
            data['users_list'].append(set)

        data['passing_score']   = conquest_data[0].passing_score
        data['completion_rate'] = c_rate

        data['pending_count']   = pending_r
        data['total_ack']       = ack_r
        data['total_pass']      = total_pass
        data['total_fail']      = total_fail

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Conquest List with Filtered Data ---------------------------------------------------
@login_required
def conquest_filtered_list(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        user_type = request.user.email

        category = request.POST.get('category_name')
        subcategory = request.POST.get('subcategory_name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        pkt_list = []

        if start_date == "na":

            if category == "Account":
                proj = []
                if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                    proj = projects_list.objects.filter(account_name = subcategory).values('project_name')
                elif user_type == "poc":
                    proj = user_details.objects.filter(email_id=request.user.username, account_name = subcategory) \
                        .values('project_name').distinct()
                else:
                    pass

                if proj != []:
                    for j in proj:
                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')
                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                sp_details = sub_process_list.objects.filter(sub_process_name = k['sub_process_name'])
                                kc_list = conquest_list.objects.filter(sub_process = k['sub_process_name'])
                                for kc in kc_list:

                                    total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                                    ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                                    if ack_r == 0:
                                        c_rate = 0.0
                                    else:
                                        c_rate  = round(float((ack_r / total_r) * 100), 2)

                                    set = {}
                                    set['timestamp']        = kc.timestamp
                                    set['id']               = kc.id
                                    set['title']            = kc.title
                                    set['sender']           = kc.sender
                                    set['account']          = sp_details[0].account_name
                                    set['project']          = sp_details[0].project_name
                                    set['sub_process']      = kc.sub_process
                                    set['questions_count']  = kc.questions_count
                                    set['passing_score']    = kc.passing_score
                                    set['max_attempts']     = kc.max_attempts
                                    set['deadline']         = kc.deadline
                                    set['completion_rate']  = c_rate
                                    pkt_list.append(set)

            if category == "Project":
                sp = []
                if user_type != "agent":
                    sp = sub_process_list.objects.filter(project_name = subcategory).values('sub_process_name')
                else:
                    pass

                if sp != []:
                    for k in sp:
                        sp_details = sub_process_list.objects.filter(sub_process_name=k['sub_process_name'])
                        kc_list = conquest_list.objects.filter(sub_process=k['sub_process_name'])
                        for kc in kc_list:

                            total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                            ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                            if ack_r == 0:
                                c_rate = 0.0
                            else:
                                c_rate  = round(float((ack_r / total_r) * 100), 2)

                            set = {}
                            set['timestamp']        = kc.timestamp
                            set['id']               = kc.id
                            set['title']            = kc.title
                            set['sender']           = kc.sender
                            set['account']          = sp_details[0].account_name
                            set['project']          = sp_details[0].project_name
                            set['sub_process']      = kc.sub_process
                            set['questions_count']  = kc.questions_count
                            set['passing_score']    = kc.passing_score
                            set['max_attempts']     = kc.max_attempts
                            set['deadline']     = kc.deadline
                            set['completion_rate']  = c_rate
                            pkt_list.append(set)

            if category == "Sub Process":
                sp_details = sub_process_list.objects.filter(sub_process_name = subcategory)
                kc_list = conquest_list.objects.filter(sub_process = subcategory)
                for kc in kc_list:

                    total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                    ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                    if ack_r == 0:
                        c_rate = 0.0
                    else:
                        c_rate  = round(float((ack_r / total_r) * 100), 2)

                    set = {}
                    set['timestamp']        = kc.timestamp
                    set['id']               = kc.id
                    set['title']            = kc.title
                    set['sender']           = kc.sender
                    set['account']          = sp_details[0].account_name
                    set['project']          = sp_details[0].project_name
                    set['sub_process']      = kc.sub_process
                    set['questions_count']  = kc.questions_count
                    set['passing_score']    = kc.passing_score
                    set['max_attempts']     = kc.max_attempts
                    set['deadline']     = kc.deadline
                    set['completion_rate']  = c_rate
                    pkt_list.append(set)

        else:

            if category == "Account":
                proj = []
                if user_type == "superuser" or user_type == "leader" or user_type == "admin":
                    proj = projects_list.objects.filter(account_name=subcategory).values('project_name')
                elif user_type == "poc":
                    proj = user_details.objects.filter(email_id=request.user.username, account_name=subcategory) \
                        .values('project_name').distinct()
                else:
                    pass

                if proj != []:
                    for j in proj:
                        sp = []
                        if user_type != "agent":
                            sp = sub_process_list.objects.filter(project_name=j['project_name']).values(
                                'sub_process_name')
                        else:
                            pass

                        if sp != []:
                            for k in sp:
                                sp_details = sub_process_list.objects.filter(sub_process_name=k['sub_process_name'])
                                kc_list = conquest_list.objects.filter(sub_process=k['sub_process_name'])
                                for kc in kc_list:
                                    val1 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                                        start_date, "%d-%m-%Y")
                                    val2 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                                      "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                                "%d-%m-%Y")
                                    if val1 and val2:

                                        total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                                        ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                                        if ack_r == 0:
                                            c_rate = 0.0
                                        else:
                                            c_rate  = round(float((ack_r / total_r) * 100), 2)

                                        set = {}
                                        set['timestamp']        = kc.timestamp
                                        set['id']               = kc.id
                                        set['title']            = kc.title
                                        set['sender']           = kc.sender
                                        set['account']          = sp_details[0].account_name
                                        set['project']          = sp_details[0].project_name
                                        set['sub_process']      = kc.sub_process
                                        set['questions_count']  = kc.questions_count
                                        set['passing_score']    = kc.passing_score
                                        set['max_attempts']     = kc.max_attempts
                                        set['deadline']         = kc.deadline
                                        set['completion_rate']  = c_rate
                                        pkt_list.append(set)

            if category == "Project":
                sp = []
                if user_type != "agent":
                    sp = sub_process_list.objects.filter(project_name=subcategory).values('sub_process_name')
                else:
                    pass

                if sp != []:
                    for k in sp:
                        sp_details = sub_process_list.objects.filter(sub_process_name=k['sub_process_name'])
                        kc_list = conquest_list.objects.filter(sub_process=k['sub_process_name'])
                        for kc in kc_list:
                            val1 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                              "%d-%m-%Y") >= datetime.datetime.strptime(
                                start_date, "%d-%m-%Y")
                            val2 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                              "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                        "%d-%m-%Y")
                            if val1 and val2:

                                total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                                ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                                if ack_r == 0:
                                    c_rate = 0.0
                                else:
                                    c_rate  = round(float((ack_r / total_r) * 100), 2)

                                set = {}
                                set['timestamp']        = kc.timestamp
                                set['id']               = kc.id
                                set['title']            = kc.title
                                set['sender']           = kc.sender
                                set['account']          = sp_details[0].account_name
                                set['project']          = sp_details[0].project_name
                                set['sub_process']      = kc.sub_process
                                set['questions_count']  = kc.questions_count
                                set['passing_score']    = kc.passing_score
                                set['max_attempts']     = kc.max_attempts
                                set['deadline']         = kc.deadline
                                set['completion_rate']  = c_rate
                                pkt_list.append(set)

            if category == "Sub Process":
                sp_details = sub_process_list.objects.filter(sub_process_name = subcategory)
                kc_list = conquest_list.objects.filter(sub_process = subcategory)
                for kc in kc_list:
                    val1 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                      "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date, "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(kc.timestamp[0:10],
                                                      "%d-%m-%Y") <= datetime.datetime.strptime(end_date,
                                                                                                "%d-%m-%Y")
                    if val1 and val2:

                        total_r = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                        ack_r   = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                        if ack_r == 0:
                            c_rate = 0.0
                        else:
                            c_rate  = round(float((ack_r / total_r) * 100), 2)

                        set = {}
                        set['timestamp']        = kc.timestamp
                        set['id']               = kc.id
                        set['title']            = kc.title
                        set['sender']           = kc.sender
                        set['account']          = sp_details[0].account_name
                        set['project']          = sp_details[0].project_name
                        set['sub_process']      = kc.sub_process
                        set['questions_count']  = kc.questions_count
                        set['passing_score']    = kc.passing_score
                        set['max_attempts']     = kc.max_attempts
                        set['deadline']         = kc.deadline
                        set['completion_rate']  = c_rate
                        pkt_list.append(set)

        data['conquest_details'] = pkt_list

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Change Knowledge Test validity -----------------------------------------------------
@login_required
def change_kc_validity(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        kc_list = json.loads(request.POST['kc_list'])
        validity = request.POST.get('validity')

        kc_count = len(kc_list)

        for i in range(kc_count):
            conquest_list.objects.filter(pk = kc_list[i]).update(deadline = validity)


        data['message'] = "Validity Changed Successfully"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Conquest -----------------------------------------------------------------------





# --------------------------------- QUESTION BANK ----------------------------------------------------------------------


# --------------------------------- REDIRECT TO QUESTION BANK ----------------------------------------------------------
@login_required
def question_bank(request):
    request.session.set_expiry(3600)
    usrname = request.user.username
    user_type = request.user.email

    if user_type == "superuser" or user_type == "leader":
        project_name = projects_list.objects.all().values('project_name').distinct()

    elif user_type == "admin":
        project_name = []
        acc_list = user_details.objects.filter(email_id = usrname).values('account_name').distinct()
        for ac in acc_list:
            project = projects_list.objects.filter(account_name = ac['account_name']).values('project_name').distinct()
            project_name.append(project)
        project_name = project_name[0]

    else:
        project_name = user_details.objects.filter(email_id = usrname).values('project_name').distinct()

    sub_process = ""
    update_details_list = ""
    update_id_list = ""
    for x in project_name:
        grp = sub_process_list.objects.filter(project_name=x['project_name'])
        sub_process = list(chain(sub_process, grp))

    for x in sub_process:
        update_details_list = list(
            chain(update_details_list, update_details.objects.filter(sub_process=x.sub_process_name).order_by('-id')))
        
    for x in update_details_list:
        if update_conquest.objects.filter(update_id=x.id):
            pass
        else:
            current_timestamp = datetime.datetime.now()
            deadline_timestamp = datetime.datetime.strptime(x.deadline, "%d-%m-%Y %H:%M")

            if current_timestamp < deadline_timestamp:
                update_id_list = list(chain(update_id_list, update_details.objects.filter(id=x.id)))

    creator_list = []
    cr_list = []
    for c in question_list.objects.all().values('creator').distinct():
        for x in project_name:
            pr_check = user_details.objects.filter(email_id = c['creator'], project_name = x['project_name']).count()
            if pr_check > 0:
                cr_list.append(c['creator'])

    for x in cr_list:
        if x not in creator_list:
            creator_list.append(x)

    questions = []
    for c in creator_list:
        q_list = question_list.objects.filter(creator = c)
        for q in q_list:

            correct_answer = []
            ques_options = question_details.objects.filter(question_id = q.id)
            op_count = 0
            for op in ques_options:
                op_count = op_count + 1
                if q.type != "Subjective":
                    if op.question_answer == "yes":
                        correct_answer.append(op_count)
                else:
                    correct_answer.append(op.question_answer)

            total_recipients = conquest_recipients_answers.objects.filter(question_id = q.id).count()
            total_pass_recipients = conquest_recipients_answers.objects.filter(question_id= q.id, answers=correct_answer).count()
            if total_recipients != 0 and total_pass_recipients != 0:
                total_pass_percentage = round((total_pass_recipients / total_recipients) * 100, 2)
            else:
                total_pass_percentage = 0.0

            object = {}
            object['id']                    = q.id
            object['timestamp']             = q.timestamp
            object['creator']               = q.creator
            object['question']              = q.question
            object['type']                  = q.type
            object['passing_percentage']    = total_pass_percentage
            questions.append(object)

    data = {}
    data['questions'] = questions
    data['sub_process'] = sub_process
    data['update_id'] = update_id_list
    return render(request, 'd9-question-bank.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SEND KNOWLEDGE CHECK FROM QUESTION BANK --------------------------------------------
def send_kc_ques(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        up_id = request.POST.get('up_id')
        timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        sub_process = json.loads(request.POST['sub_process'])
        validity = request.POST.get('validity')
        score = request.POST.get('score')
        attempts = request.POST.get('attempts')
        kc_title = request.POST.get('kc_title')
        questions = json.loads(request.POST['questions'])
        q_count = len(questions)
        sender = request.user.username

        if up_id != "na":
            up_details = update_details.objects.filter(id=up_id)
            subprocess = up_details[0].sub_process
            validity = up_details[0].deadline

            pkt_id = save_kc_details(timestamp, subprocess, q_count, score, sender, validity, attempts, kc_title)

            for q_id in questions:
                conq_ques = conquest_questions(conquest_id=pkt_id, question_id=q_id)
                conq_ques.save()

            sbp_users_list = sub_process_details.objects.filter(sub_process_name=subprocess)

            receiver = []
            for usr in sbp_users_list:
                conq_id = pkt_id
                agent = usr.user_email
                q_attempts = "0"
                q_score = "0"
                q_status = "na"
                t_stamp = "NA"
                save_conquest_recipients(conq_id, agent, q_attempts, q_score, q_status, t_stamp)
                receiver.append(agent)

            save_up_conq(up_id, pkt_id)

            cc_receiver = cc_list_of_poc_and_admins(sub_process, sender)
            
            data['message'] = "Knowledge Test Sent successfully for a previously sent Update"
            send_update_pkt_email(sender, receiver, cc_receiver)

        else:
            sub_process_count = len(sub_process)
            for i in range(sub_process_count):

                pkt_id = save_kc_details(timestamp, sub_process[i], q_count, score, sender, validity, attempts, kc_title)

                for q_id in questions:
                    conq_ques = conquest_questions(conquest_id=pkt_id, question_id=q_id)
                    conq_ques.save()

                sbp_users_list = sub_process_details.objects.filter(sub_process_name=sub_process[i])

                receiver = []
                for usr in sbp_users_list:
                    conq_id = pkt_id
                    agent = usr.user_email
                    q_attempts = "0"
                    q_score = "0"
                    q_status = "na"
                    t_stamp = "NA"
                    save_conquest_recipients(conq_id, agent, q_attempts, q_score, q_status, t_stamp)
                    receiver.append(agent)

                cc_receiver = cc_list_of_poc_and_admins(sub_process[i], sender)

                send_pkt_email(sender, receiver, cc_receiver)

            data['message'] = "Knowledge Test Sent successfully"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DETAILS OF SELECTED QUESTION FRM QUESTION BANK -------------------------------------
@login_required
def question_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        question_id             = request.POST['question_id']
        ques_details            = question_list.objects.filter(id=question_id)
        data['question']        = ques_details[0].question
        data['timestamp']       = ques_details[0].timestamp
        data['question_type']   = ques_details[0].type

        option = []
        options = question_details.objects.filter(question_id=question_id)

        correct_answer = []
        op_count = 0
        for x in options:
            set = {}
            set['option'] = x.question_option
            set['answer'] = x.question_answer
            option.append(set)

            op_count = op_count + 1
            if ques_details[0].type != "Subjective":
                if x.question_answer == "yes":
                    correct_answer.append(op_count)
            else:
                correct_answer.append(x.question_answer)

        data['options'] = option

        users_list = conquest_recipients_answers.objects.filter(question_id=question_id)
        data['users_list'] = []

        for usr in users_list:

            if user_details.objects.filter(email_id=usr.user_email):
                emp_id = user_details.objects.filter(email_id=usr.user_email)[0].emp_id
            else:
                emp_id = "Removed"

            if User.objects.filter(username=usr.user_email).values('first_name'):
                emp_f_name = User.objects.filter(username=usr.user_email).values('first_name')[0]['first_name']
            else:
                emp_f_name = "Removed"

            if User.objects.filter(username=usr.user_email).values('last_name'):
                emp_l_name = User.objects.filter(username=usr.user_email).values('last_name')[0]['last_name']
            else:
                emp_l_name = "Removed"

            if emp_f_name == "Removed":
                emp_name = "Removed"
            else:
                emp_name = emp_f_name + " " + emp_l_name

            answer = eval(usr.answers)
            answer_length = len(answer)
            q_ans = ""

            if ques_details[0].type == "Subjective":
                q_ans = q_ans + answer[0]

            elif ques_details[0].type == "True or False" or ques_details[0].type == "MCQs":
                for a_op in range(answer_length):
                    val = int(answer[a_op])
                    q_ans = q_ans + options[val - 1].question_option

            else:
                for a_op in range(answer_length):
                    val = int(answer[a_op])
                    q_ans = q_ans + options[val - 1].question_option + ", "

            set = {}
            set['emp_id']           = emp_id
            set['emp_name']         = emp_name
            set['emp_email']        = usr.user_email
            set['emp_answer']       = q_ans
            if str(correct_answer) == usr.answers:
                set['emp_status']   = "Correct"
            else:
                set['emp_status']   = "Incorrect"
            data['users_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Create Questions BY UPLOADING EXCEL-SHEET ------------------------------------------
def upload_questions_file(request):
    if request.method == 'POST' and request.FILES['questions_file']:
        request.session.set_expiry(3600)
        myfile = request.FILES['questions_file']
        fs = FileSystemStorage('media/excel_uploads/question_bank/')
        filename = fs.save((myfile.name).replace(" ", "_"), myfile)
        creator = request.user.username
        uploaded_file_url = fs.url("/excel_uploads/question_bank/" + filename)
        create_questions_from_excel(uploaded_file_url, creator)

    return redirect('/dashboard/question_bank/')

def create_questions_from_excel(file_path, creator):
    book = openpyxl.load_workbook(str(settings.BASE_DIR) + file_path)
    sheet = book.active

    for r in range(1, sheet.max_row):
        type        = (sheet.cell(row =r, column = 0).value).strip()
        question    = (sheet.cell(row =r,  column =1).value).strip()
        marks       = int((sheet.cell(row =r, column = 2).value))

        timestamp   = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        media_type  = "No Media"
        media_url   = ""
        question_id     = save_question_fn(question, type, timestamp, media_type, media_url, marks, creator)

        op_count    = int((sheet.cell(row =r,  column =3).value))
        counter = 0
        for op in range(op_count):
            question_option = str(sheet.cell(row =r, column = (4 + counter)).value).strip()
            question_answer = str(sheet.cell(row =r, column = (5 + counter)).value).strip()
            save_option_fn(question_id, question_option, question_answer)
            counter = counter + 2
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DOWNLOAD BULK-UPLOAD IN Questions TEMPLATE -----------------------------------------

def download_questions_template(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')

    if csrf_token != None:
        request.session.set_expiry(3600)
        path = '/excel_upload_templates/upload_questions.xls'
        file_path = str(settings.MEDIA_ROOT) + path
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END QUESTION BANK ------------------------------------------------------------------


# --------------------------------- Conquest Response ------------------------------------------------------------------


# --------------------------------- REDIRECT TO CONQUEST RESPONSE ------------------------------------------------------
@login_required
def conquest_response(request):
    request.session.set_expiry(3600)
    usrname = request.user.username
    user_type = request.user.email

    if user_type == "superuser" or user_type == "leader":
        project_name = projects_list.objects.all().values('project_name').distinct()

    elif user_type == "admin":
        project_name = []
        acc_list = user_details.objects.filter(email_id=usrname).values('account_name').distinct()
        for ac in acc_list:
            project = projects_list.objects.filter(account_name=ac['account_name']).values('project_name').distinct()
            project_name.append(project)
        project_name = project_name[0]

    else:
        project_name = user_details.objects.filter(email_id=usrname).values('project_name').distinct()

    sp_list = []
    for x in project_name:
        sp = sub_process_list.objects.filter(project_name=x['project_name']).values('sub_process_name')
        for y in sp:
            sp_set = {}
            sp_set['sub_process_name'] = y['sub_process_name']
            sp_list.append(sp_set)

    conquest = []
    for x in sp_list:
        kc_list = conquest_list.objects.filter(sub_process=x['sub_process_name'])

        for y in kc_list:
            object = {}
            object['id'] = y.id
            object['title'] = y.title
            conquest.append(object)

    data = {}
    data['conquest'] = conquest
    return render(request, 'd8-conquest-response.html', data)


# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Details of CONQUEST RESPONSES As per Knowledge Check Selected ----------------------
def conquest_responses_data(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        title = request.POST.get('conquest_id')
        conquest_data = conquest_list.objects.filter(title=title)
        conquest_id = conquest_data[0].id

        responses = []

        conquest_data = conquest_list.objects.filter(id=conquest_id)

        # Table heading
        set = {}
        set['emp_id'] = "Emp Id"
        set['emp_name'] = "Emp Name"
        set['emp_email'] = "Emp Email"
        set['timestamp'] = "Submitted At"
        set['sub_process'] = "SubProcess"
        set['created_on'] = "Created On"
        set['conquest_id'] = "KC Title"
        set['attempt'] = "Total Attempts"
        set['questions_list'] = []
        ques_list = conquest_questions.objects.filter(conquest_id=conquest_id)
        i = 1
        for q in ques_list:
            q_count = "Question-" + str(i)
            set['questions_list'].append(q_count)
            i = i + 1
        set['score'] = "Scored Percentage"
        set['status'] = "Status"
        responses.append(set)
        # End of Table Heading

        # Table First row - List of Questions of Knowledge Check
        set = {}
        set['emp_id'] = "-"
        set['emp_name'] = "-"
        set['emp_email'] = "-"
        set['timestamp'] = "-"
        set['sub_process'] = conquest_data[0].sub_process
        set['created_on'] = conquest_data[0].timestamp
        set['conquest_id'] = title
        set['attempt'] = "-"
        set['questions_list'] = []
        ques_list = conquest_questions.objects.filter(conquest_id=conquest_id)
        for q in ques_list:
            q_detail = question_list.objects.filter(id=q.question_id)
            set['questions_list'].append(q_detail[0].question)
        set['score'] = conquest_data[0].passing_score + "%"
        set['status'] = "-"
        responses.append(set)
        # End of Table First row - List of Questions of Knowledge Check

        # Table Second row - Question's Answers of Knowledge Check
        set = {}
        set['emp_id'] = "-"
        set['emp_name'] = "-"
        set['emp_email'] = "-"
        set['timestamp'] = "-"
        set['sub_process'] = conquest_data[0].sub_process
        set['created_on'] = conquest_data[0].timestamp
        set['conquest_id'] = title
        set['attempt'] = "-"
        set['questions_list'] = []
        ques_list = conquest_questions.objects.filter(conquest_id=conquest_id)
        for q in ques_list:
            q_detail = question_list.objects.filter(id=q.question_id)
            ans_list = question_details.objects.filter(question_id=q.question_id)
            q_ans = ""
            if q_detail[0].type == "Subjective":
                for qa in ans_list:
                    q_ans = q_ans + qa.question_answer + ", "

            elif q_detail[0].type == "True or False" or q_detail[0].type == "MCQs":
                for qa in ans_list:
                    if qa.question_answer == "yes":
                        q_ans = q_ans + qa.question_option

            else:
                for qa in ans_list:
                    if qa.question_answer == "yes":
                        q_ans = q_ans + qa.question_option + ", "
            set['questions_list'].append(q_ans)
        set['score'] = "-"
        set['status'] = "-"
        responses.append(set)
        # End of Table Second row - Question's Answers of Knowledge Check

        # Table Recipients Responses
        recipients_list = conquest_recipients.objects.filter(conquest_id=conquest_id)

        for r in recipients_list:
            if int(r.attempts) > 0:

                recipient_details = user_details.objects.filter(email_id=r.user_email)

                if recipient_details.exists():
                    emp_id = recipient_details[0].emp_id
                    emp_name = recipient_details[0].full_name
                    emp_email = recipient_details[0].email_id
                else:
                    emp_id = "Removed"
                    emp_name = "Removed"
                    emp_email = "Removed"

                attempt_no = int(r.attempts)
                set = {}
                set['emp_id'] = emp_id
                set['emp_name'] = emp_name
                set['emp_email'] = emp_email
                set['timestamp'] = r.timestamp
                set['sub_process'] = conquest_data[0].sub_process
                set['created_on'] = conquest_data[0].timestamp
                set['conquest_id'] = title
                set['attempt'] = attempt_no
                set['questions_list'] = []
                ques_list = conquest_questions.objects.filter(conquest_id=conquest_id)
                for q in ques_list:
                    q_detail = question_list.objects.filter(id=q.question_id)
                    ans_details = question_details.objects.filter(question_id=q.question_id)
                    ans_list = conquest_recipients_answers.objects.filter(conquest_id=conquest_id,
                                                                          question_id=q.question_id,
                                                                          user_email=r.user_email,
                                                                          attempt_no=attempt_no).values('answers')
                    answer = eval(ans_list[0]['answers'])
                    answer_length = len(answer)
                    q_ans = ""
                    if q_detail[0].type == "Subjective":
                        q_ans = q_ans + answer[0]

                    elif q_detail[0].type == "True or False" or q_detail[0].type == "MCQs":
                        for a_op in range(answer_length):
                            val = int(answer[a_op])
                            q_ans = q_ans + ans_details[val - 1].question_option

                    else:
                        for a_op in range(answer_length):
                            val = int(answer[a_op])
                            q_ans = q_ans + ans_details[val - 1].question_option + ", "

                    set['questions_list'].append(q_ans)
                set['score'] = str(round(float(r.score), 2)) + "%"
                if float(r.score) < float(conquest_data[0].passing_score):
                    set['status'] = "Failed"
                else:
                    set['status'] = "Passed"
                responses.append(set)
        # Table Recipients Responses

        data['conquest_response'] = responses

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)


# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Conquest Response --------------------------------------------------------------





# --------------------------------- Create Knowledge Check -------------------------------------------------------------


# --------------------------------- REDIRECT TO Create Knowledge Check -------------------------------------------------
@login_required
def createKC(request):
    request.session.set_expiry(3600)
    proj_list = user_details.objects.filter(email_id=request.user.username).values('project_name').distinct()
    sub_process = ""
    update_details_list = ""
    update_id_list = ""
    for x in proj_list:
        grp = sub_process_list.objects.filter(project_name=x['project_name'])
        sub_process = list(chain(sub_process, grp))

    for x in sub_process:
        update_details_list = list(
            chain(update_details_list, update_details.objects.filter(sub_process=x.sub_process_name).order_by('-id')))

    for x in update_details_list:
        if update_conquest.objects.filter(update_id=x.id):
            pass
        else:
            current_timestamp = datetime.datetime.now()
            deadline_timestamp = datetime.datetime.strptime(x.deadline, "%d-%m-%Y %H:%M")

            if current_timestamp < deadline_timestamp:
                update_id_list = list(chain(update_id_list, update_details.objects.filter(id=x.id)))

    data = {}
    data['sub_process'] = sub_process
    data['update_id'] = update_id_list
    return render(request, 'd2-create-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Create Knowledge Check ---------------------------------------------------------





# --------------------------------- Create A Question ------------------------------------------------------------------


# --------------------------------- REDIRECT TO Create A Question ------------------------------------------------------
@login_required
def createQuestion(request):
    request.session.set_expiry(3600)
    proj_list = user_details.objects.filter(email_id=request.user.username).values('project_name').distinct()
    sub_process = ""
    conquest_details_list = ""
    conquest_id_list = ""

    for x in proj_list:
        grp = sub_process_list.objects.filter(project_name=x['project_name'])
        sub_process = list(chain(sub_process, grp))

    for x in sub_process:
        conquest_details_list = list(
            chain(conquest_details_list, conquest_list.objects.filter(sub_process=x.sub_process_name).order_by('-id')))

    for x in conquest_details_list:
        current_timestamp   = datetime.datetime.now()
        deadline_timestamp  = datetime.datetime.strptime(x.deadline, "%d-%m-%Y %H:%M")

        total_r             = conquest_recipients.objects.filter(conquest_id=x.id).count()
        ack_r               = conquest_recipients.objects.filter(conquest_id=x.id).exclude(attempts="0").count()

        if ack_r == 0:
            c_rate          = 0.0
        else:
            c_rate          = round(float((ack_r / total_r) * 100), 2)

        if current_timestamp < deadline_timestamp:
            if c_rate == 0:
                conquest_id_list = list(chain(conquest_id_list, conquest_list.objects.filter(id=x.id)))

    data = {}
    data['question_editor'] = conquest_questionForm()
    data['conquest_id'] = conquest_id_list

    return render(request, 'd3-create-a-question.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE THE QUESTION DETAILS PROVIDED -------------------------------------------------
@login_required
def save_question(request):
    request.session.set_expiry(3600)
    timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    question = request.POST['question']
    type = request.POST['type']
    media_type = request.POST['media_type']
    media_url = request.POST['media_url']
    marks = request.POST['marks']
    creator = request.user.username

    ques_id = save_question_fn(question, type, timestamp, media_type, media_url, marks, creator)

    return JsonResponse(ques_id, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE THE OPTIONS OF QUESTION PROVIDED ----------------------------------------------
@login_required
def save_option(request):
    request.session.set_expiry(3600)
    question_id = request.POST['question_id']
    question_option = request.POST['question_option']
    question_answer = request.POST['question_answer']

    save_option_fn(question_id, question_option, question_answer)

    return render(request, 'd3-create-a-question.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE THE Question with Knowledge Check ---------------------------------------------
@login_required
def save_conquest_question(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        timestamp = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        title = request.POST.get('conquest_id')
        conquest_data = conquest_list.objects.filter(title=title)
        pkt_id = conquest_data[0].id

        questions = json.loads(request.POST['questions'])

        q_count = conquest_list.objects.filter(pk=pkt_id)
        q_count = int(q_count[0].questions_count) + 1
        conquest_list.objects.filter(pk=pkt_id).update(questions_count=q_count)

        for a in questions:
            question = a['question']
            type = a['question_type']
            media_type = a['question_media_type']
            media_url = a['question_media_url']
            marks = a['question_marks']
            creator =  request.user.username

            q_id = save_question_fn(question, type, timestamp, media_type, media_url, marks, creator)

            conq_ques = conquest_questions(conquest_id=pkt_id,
                                           question_id=q_id)
            conq_ques.save()

            op_c = 0
            for op in a['question_options']:
                ans_c = 0
                for ans in a['question_answer']:
                    if op_c == ans_c:
                        question_option = op
                        q_ans_val = ans
                        if a['question_type'] != "Subjective":
                            if q_ans_val == 0:
                                question_answer = "no"
                            else:
                                question_answer = "yes"
                        else:
                            question_answer = q_ans_val
                        save_option_fn(q_id, question_option, question_answer)
                    ans_c = ans_c + 1
                op_c = op_c + 1

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- REGISTER USERS BY UPLOADING EXCEL-SHEET --------------------------------------------
def upload_media_file(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        acc_name = user_details.objects.filter(email_id=request.user.username).values('account_name')

        myfile = request.FILES['user_file']
        file_ext = myfile.name.split(".")[-1]

        if file_ext.lower() == "mp4" or \
            file_ext.lower() == "avi" or \
            file_ext.lower() == "mp3" or \
            file_ext.lower() == "wav" or \
            file_ext.lower() == "xls" or \
            file_ext.lower() == "xlsx" or \
            file_ext.lower() == "docx" or \
            file_ext.lower() == "doc" or \
            file_ext.lower() == "pdf" :

            fs = FileSystemStorage('media/media_upload/' + acc_name[0]['account_name'].replace(" ", "_") + '/')
            filename = fs.save((myfile.name).replace(" ", "_"), myfile)
            uploaded_file_url = fs.url("/media_upload/" + acc_name[0]['account_name'].replace(" ", "_") + "/" + filename)


            data['code'] = 1
            data['message'] = "File uploaded successfully"
            data['uploaded_media_url'] = uploaded_file_url

        else:
            data['code'] = 0
            data['message'] = "Upload Unsuccessful : Invalid File Type"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- REGISTER USERS BY UPLOADING EXCEL-SHEET --------------------------------------------


# --------------------------------- END Create A Question --------------------------------------------------------------





# --------------------------------- All Knowledge Check ----------------------------------------------------------------


# --------------------------------- REDIRECT TO All Knowledge Check ----------------------------------------------------
@login_required
def all_KC(request):
    request.session.set_expiry(3600)
    account_list = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

    sp_list = ""
    for x in account_list:
        sp = sub_process_list.objects.filter(account_name=x['account_name'])
        sp_list = list(chain(sp_list, sp))

    kc_id_list = []
    for x in sp_list:
        kc_id = conquest_list.objects.filter(sub_process=x.sub_process_name).order_by('-id')
        for id in kc_id:
            kc_id_list.append(id.id)

    kc_id_list.sort(reverse = True)

    kc_list = ""
    for id in kc_id_list:
        kc_up_check = update_conquest.objects.filter(conquest_id=id).count()
        if kc_up_check == 0:
            kcd = conquest_list.objects.filter(pk=id)
            kc_list = list(chain(kc_list, kcd))

    data = {}
    data['conquest_details'] = kc_list
    return render(request, 'd4-all-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER KNOWLEDGE CHECK AS PER DATE RANGE SELECTED ----------------------------------
@login_required
def allkc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        account_list = user_details.objects.filter(email_id=request.user.username).values('account_name').distinct()

        sp_list = ""
        for x in account_list:
            sp = sub_process_list.objects.filter(account_name=x['account_name'])
            sp_list = list(chain(sp_list, sp))

        kc_id_list = []
        for x in sp_list:
            kc_id = conquest_list.objects.filter(sub_process=x.sub_process_name).order_by('-id')
            for id in kc_id:
                kc_id_list.append(id.id)

        kc_id_list.sort(reverse=True)

        data['conquest_details'] = []
        for id in kc_id_list:
            kc_up_check = update_conquest.objects.filter(conquest_id= str(id)).count()
            if kc_up_check == 0:
                kcd = conquest_list.objects.filter(pk=id)
                for y in kcd:
                    val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date,
                        "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                        end_date,
                        "%d-%m-%Y")
                    if val1:
                        if val2:

                            for kc in kcd:

                                total_r     = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                                ack_r       = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                                if ack_r == 0:
                                    c_rate  = 0.0
                                else:
                                    c_rate  = round(float((ack_r / total_r) * 100), 2)

                                set = {}
                                set['id']               = kc.id
                                set['timestamp']        = kc.timestamp
                                set['sub_process']      = kc.sub_process
                                set['questions_count']  = kc.questions_count
                                set['passing_score']    = kc.passing_score
                                set['max_attempts']     = kc.max_attempts
                                set['sender']           = kc.sender
                                set['deadline']         = kc.deadline
                                set['completion_rate']  = c_rate
                                data['conquest_details'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END All Knowledge Check ------------------------------------------------------------





# --------------------------------- Sent Knowledge Check ---------------------------------------------------------------


# --------------------------------- REDIRECT TO Sent Knowledge Check ---------------------------------------------------
@login_required
def sent_KC(request):
    request.session.set_expiry(3600)
    kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by('-id')

    kc_list = ""
    for x in kc_id_list:
        kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
        if kc_up_check == 0:
            kcd = conquest_list.objects.filter(pk=x.id)
            kc_list = list(chain(kc_list, kcd))

    data = {}
    data['conquest_details'] = kc_list
    return render(request, 'd5-sent-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER KNOWLEDGE CHECK AS PER DATE RANGE SELECTED ----------------------------------
@login_required
def sentkc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by('-id')

        data['conquest_details'] = []
        for x in kc_id_list:
            kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
            if kc_up_check == 0:
                kcd = conquest_list.objects.filter(pk=x.id)
                for y in kcd:
                    val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                        start_date,
                        "%d-%m-%Y")
                    val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                        end_date,
                        "%d-%m-%Y")
                    if val1:
                        if val2:

                            for kc in kcd:

                                total_r     = conquest_recipients.objects.filter(conquest_id = kc.id).count()
                                ack_r       = conquest_recipients.objects.filter(conquest_id = kc.id).exclude(attempts="0").count()

                                if ack_r == 0:
                                    c_rate  = 0.0
                                else:
                                    c_rate  = round(float((ack_r / total_r) * 100), 2)

                                set = {}
                                set['id']               = kc.id
                                set['timestamp']        = kc.timestamp
                                set['sub_process']      = kc.sub_process
                                set['questions_count']  = kc.questions_count
                                set['passing_score']    = kc.passing_score
                                set['max_attempts']     = kc.max_attempts
                                set['sender']           = kc.sender
                                set['deadline']         = kc.deadline
                                set['completion_rate']  = c_rate
                                data['conquest_details'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Sent Knowledge Check -----------------------------------------------------------





# --------------------------------- Pending Knowledge Check ------------------------------------------------------------


# --------------------------------- REDIRECT TO Pending Knowledge Check ------------------------------------------------
@login_required
def pending_KC(request):
    request.session.set_expiry(3600)
    kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by(
        '-id')

    kc_list = ""
    for x in kc_id_list:

        total_r = conquest_recipients.objects.filter(conquest_id = x.id).count()
        ack_r = conquest_recipients.objects.filter(conquest_id = x.id).exclude(attempts="0").count()

        if ack_r == 0:
            c_rate = 0.0
        else:
            c_rate = round(float((ack_r / total_r) * 100), 2)

        if c_rate != 100.0:
            kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
            if kc_up_check == 0:
                kcd = conquest_list.objects.filter(pk=x.id)
                kc_list = list(chain(kc_list, kcd))

    data = {}
    data['conquest_details'] = kc_list
    return render(request, 'd6-pending-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER KNOWLEDGE CHECK AS PER DATE RANGE SELECTED ----------------------------------
@login_required
def pendingkc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by('-id')

        data['conquest_details'] = []
        for x in kc_id_list:

            total_r = conquest_recipients.objects.filter(conquest_id = x.id).count()
            ack_r = conquest_recipients.objects.filter(conquest_id = x.id).exclude(attempts="0").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate = round(float((ack_r / total_r) * 100), 2)

            if c_rate != 100.0:
                kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
                if kc_up_check == 0:
                    kcd = conquest_list.objects.filter(pk=x.id)
                    for y in kcd:
                        val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                            start_date,
                            "%d-%m-%Y")
                        val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                            end_date,
                            "%d-%m-%Y")
                        if val1:
                            if val2:

                                for kc in kcd:

                                    total_r = conquest_recipients.objects.filter(conquest_id=kc.id).count()
                                    ack_r = conquest_recipients.objects.filter(conquest_id=kc.id).exclude(
                                        attempts="0").count()

                                    if ack_r == 0:
                                        c_rate = 0.0
                                    else:
                                        c_rate = round(float((ack_r / total_r) * 100), 2)

                                    set = {}
                                    set['id'] = kc.id
                                    set['timestamp'] = kc.timestamp
                                    set['sub_process'] = kc.sub_process
                                    set['questions_count'] = kc.questions_count
                                    set['passing_score'] = kc.passing_score
                                    set['max_attempts'] = kc.max_attempts
                                    set['sender'] = kc.sender
                                    set['deadline'] = kc.deadline
                                    set['completion_rate'] = c_rate
                                    data['conquest_details'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Pending Knowledge Check --------------------------------------------------------





# --------------------------------- Completed Knowledge Check ----------------------------------------------------------


# --------------------------------- REDIRECT TO Completed Knowledge Check ----------------------------------------------
@login_required
def completed_KC(request):
    request.session.set_expiry(3600)
    kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by('-id')

    kc_list = ""
    for x in kc_id_list:

        total_r = conquest_recipients.objects.filter(conquest_id = x.id).count()
        ack_r = conquest_recipients.objects.filter(conquest_id = x.id).exclude(attempts="0").count()

        if ack_r == 0:
            c_rate = 0.0
        else:
            c_rate = round(float((ack_r / total_r) * 100), 2)

        if c_rate == 100.0:
            kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
            if kc_up_check == 0:
                kcd = conquest_list.objects.filter(pk=x.id)
                kc_list = list(chain(kc_list, kcd))

    data = {}
    data['conquest_details'] = kc_list
    return render(request, 'd7-completed-knowledge-check.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- FILTER KNOWLEDGE CHECK AS PER DATE RANGE SELECTED ----------------------------------
@login_required
def completedkc_date_filter(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        kc_id_list = conquest_list.objects.filter(sender=request.user.username).order_by('-id')

        data['conquest_details'] = []
        for x in kc_id_list:

            total_r = conquest_recipients.objects.filter(conquest_id=x.id).count()
            ack_r = conquest_recipients.objects.filter(conquest_id=x.id).exclude(attempts="0").count()

            if ack_r == 0:
                c_rate = 0.0
            else:
                c_rate = round(float((ack_r / total_r) * 100), 2)

            if c_rate == 100.0:
                kc_up_check = update_conquest.objects.filter(conquest_id=x.id).count()
                if kc_up_check == 0:
                    kcd = conquest_list.objects.filter(pk=x.id)
                    for y in kcd:
                        val1 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") >= datetime.datetime.strptime(
                            start_date,
                            "%d-%m-%Y")
                        val2 = datetime.datetime.strptime(y.timestamp[0:10], "%d-%m-%Y") <= datetime.datetime.strptime(
                            end_date,
                            "%d-%m-%Y")
                        if val1:
                            if val2:

                                for kc in kcd:

                                    total_r = conquest_recipients.objects.filter(conquest_id=kc.id).count()
                                    ack_r = conquest_recipients.objects.filter(conquest_id=kc.id).exclude(
                                        attempts="0").count()

                                    if ack_r == 0:
                                        c_rate = 0.0
                                    else:
                                        c_rate = round(float((ack_r / total_r) * 100), 2)

                                    set = {}
                                    set['id'] = kc.id
                                    set['timestamp'] = kc.timestamp
                                    set['sub_process'] = kc.sub_process
                                    set['questions_count'] = kc.questions_count
                                    set['passing_score'] = kc.passing_score
                                    set['max_attempts'] = kc.max_attempts
                                    set['sender'] = kc.sender
                                    set['deadline'] = kc.deadline
                                    set['completion_rate'] = c_rate
                                    data['conquest_details'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Completed Knowledge Check ------------------------------------------------------





# --------------------------------- Update Source ----------------------------------------------------------------------


# --------------------------------- REDIRECT To Update Source ----------------------------------------------------------
@login_required
def listUpdateSource(request):
    request.session.set_expiry(3600)
    update_source = update_source_list.objects.all().order_by('update_source_name')
    return render(request, 'e1-update_source.html', {'update_source': update_source})
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- CREATE UPDATE SOURCE ---------------------------------------------------------------
@login_required
def create_update_source(request):
    data = {}
    request.session.set_expiry(3600)
    update_source = request.POST['update_source'].strip()
    check = update_source_list.objects.filter(update_source_name = update_source).count()

    if check == 0:
        update_src = update_source_list(update_source_name=update_source)
        update_src.save()
        data['message'] = "Update Source saved Successfully"
    else:
        data['message'] = "Update Source name already exists"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE UPDATE SOURCE ---------------------------------------------------------------
@login_required
def delete_update_source(request):
    request.session.set_expiry(3600)
    update_source_id = request.POST['id']
    update_source_list.objects.get(pk=update_source_id).delete()
    return render(request, 'e1-update_source.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Update Source ------------------------------------------------------------------





# --------------------------------- Update Type ------------------------------------------------------------------------


# --------------------------------- REDIRECT TO UPDATE TYPE ------------------------------------------------------------
@login_required
def listUpdateType(request):
    request.session.set_expiry(3600)
    update_type = update_type_list.objects.all().order_by('update_type_name')
    return render(request, 'e2-update_type.html', {'update_type': update_type})
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ADD UPDATE TYPE --------------------------------------------------------------------
@login_required
def create_UpdateType(request):

    data = {}
    request.session.set_expiry(3600)
    update_type = request.POST['update_type'].strip()
    check = update_type_list.objects.filter(update_type_name=update_type).count()

    if check == 0:
        update_typ = update_type_list(update_type_name=update_type)
        update_typ.save()
        data['message'] = "Update Type saved Successfully"
    else:
        data['message'] = "Update Type name already exists"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE UPDATE TYPE -----------------------------------------------------------------
@login_required
def delete_UpdateType(request):
    request.session.set_expiry(3600)
    up_type_id = request.POST['id']
    update_type_list.objects.get(pk=up_type_id).delete()
    return render(request, 'e2-update_type.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Update Type --------------------------------------------------------------------





# --------------------------------- Users (Register Users) -------------------------------------------------------------


# --------------------------------- REDIRECT TO USERS ------------------------------------------------------------------
@login_required
def listUsers(request):
    request.session.set_expiry(3600)
    data = {}
    accounts = user_details.objects.values('account_name').filter(email_id=request.user.username).distinct()

    data['users_list'] = []
    for acc in accounts:
        user_list = user_details.objects.filter(account_name = acc['account_name'])

        for u in user_list:
            set = {}
            set['id'] = u.id
            set['emp_id'] = u.emp_id
            set['emp_name'] = u.full_name
            set['emp_email'] = u.email_id
            set['account_name'] = u.account_name
            set['project_name'] = u.project_name
            data['users_list'].append(set)

    data['accounts'] = accounts

    return render(request, 'e3-create-users.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- CREATE USER ------------------------------------------------------------------------
@login_required
def create_user(request):
    request.session.set_expiry(3600)
    ac_name = request.POST['account_name']
    proj_name = request.POST['project_name']
    emp_id = request.POST['emp_id']
    f_name = request.POST['first_name']
    l_name = request.POST['last_name']
    email = request.POST['email_id']

    sender = request.user.username

    regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'

    if account_list.objects.filter(account_name=ac_name).exists() and \
            projects_list.objects.filter(account_name=ac_name, project_name=proj_name).exists() and \
            len(emp_id) > 5 and len(emp_id) < 8 and emp_id.isdigit() and all(x.isalpha() or x.isspace() for x in f_name) and \
            all(x.isalpha() or x.isspace() for x in l_name) and \
            (re.search(regex, email)):
        user_list = []

        user = {}
        user['ac_name'] = ac_name.strip()
        user['proj_name'] = proj_name.strip()
        user['f_name'] = f_name.strip()
        user['l_name'] = l_name.strip()
        user['email'] = email.strip()
        user['emp_id'] = int(emp_id.strip())
        user_list.append(user)

        add_user(user_list, sender)

    return render(request, 'e3-create-users.html')

def add_user(user_list, sender):
    user_type = "agent"

    receiver = []

    for x in user_list:
        full_name = x['f_name'] + " " + x['l_name']
        val = check_user(x['proj_name'], x['emp_id'])

        if val == 0:
            receiver.append(x['email'])
            usr_details = user_details(account_name=x['ac_name'],
                                       project_name=x['proj_name'],
                                       emp_id=x['emp_id'],
                                       full_name=full_name,
                                       email_id=x['email'])
            usr_details.save()

            if User.objects.filter(username=x['email']).exists():
                pass
            else:
                u = User.objects.create_user(username=x['email'], password="cognizant", email=user_type,
                                             first_name=x['f_name'], last_name=x['l_name'])
                u.save()

    send_email_for_user_creation(sender, "Associate", receiver, user_type)

def check_user(proj_name, empid):
    check = user_details.objects.filter(project_name=proj_name, emp_id=empid)
    if check.count() > 0:
        val = 1
    else:
        val = 0
    return val
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE USER ------------------------------------------------------------------------
def delete_user(request):
    request.session.set_expiry(3600)

    user_id_list = json.loads(request.POST['id'])

    for user_id in user_id_list:
        user_email_id = user_details.objects.filter(pk=user_id)
        user_proj_count = user_details.objects.filter(email_id=user_email_id[0].email_id).count()

        delete_user_f_proj_sp(user_email_id[0].project_name, user_email_id[0].email_id)

        if user_proj_count == 1:
            pk_val_obj = User.objects.filter(username=user_email_id[0].email_id).values('pk')
            pk_val = pk_val_obj[0]['pk']
            User.objects.get(pk=pk_val).delete()

        user_details.objects.get(pk=user_id).delete()

    return render(request, 'e3-create-users.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE USER FROM PROJECT-SUBPROCESS ------------------------------------------------
def delete_user_f_proj_sp(project, email):
    sub_process_user_id = sub_process_details.objects. \
        filter(project_name = project, user_email = email)

    for i in sub_process_user_id:
        sub_process_details.objects.get(pk=i.id).delete()
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- Load projects in DropDown As per Account Selected ----------------------------------
def load_d_projects(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        account_name = request.POST.get('account_name')
        projects = user_details.objects.filter(email_id=request.user.username, account_name=account_name).values(
            'project_name').distinct()
        data = json.dumps(list(projects))

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- REGISTER USERS BY UPLOADING EXCEL-SHEET --------------------------------------------
def upload_users_file(request):
    if request.method == 'POST' and request.FILES['user_file']:
        request.session.set_expiry(3600)
        myfile = request.FILES['user_file']
        fs = FileSystemStorage('media/excel_uploads/users/')
        filename = fs.save((myfile.name).replace(" ", "_"), myfile)
        sender = request.user.username
        uploaded_file_url = fs.url("/excel_uploads/users/" + filename)
        create_users_from_excel(uploaded_file_url, sender)

    return redirect('/dashboard/users/')

def create_users_from_excel(file_path, sender):
    book = xlrd.open_workbook(str(settings.BASE_DIR) + file_path)
    sheet = book.sheet_by_index(0)

    user_list = []
    # loop over each row
    for r in range(1, sheet.nrows):
        # extract each cell
        user = {}
        user['ac_name']     = (sheet.cell(r, 0).value).strip()
        user['proj_name']   = (sheet.cell(r, 1).value).strip()
        user['f_name']      = (sheet.cell(r, 2).value).strip()
        user['l_name']      = (sheet.cell(r, 3).value).strip()
        user['email']       = (sheet.cell(r, 4).value).strip()
        user['emp_id']      = int((sheet.cell(r, 5).value))

        user_list.append(user)

    add_user(user_list, sender)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DOWNLOAD BULK-UPLOAD IN USERS TEMPLATE ---------------------------------------------
@login_required
def download_user_template(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')

    if csrf_token != None:
        request.session.set_expiry(3600)
        path = '/excel_upload_templates/upload_users.xlsx'
        file_path = str(settings.MEDIA_ROOT) + path
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Users (Register Users) ---------------------------------------------------------





# --------------------------------- Sub Process (Create Sub Process) ---------------------------------------------------


# --------------------------------- REDIRECT TO SUBPROCESS -------------------------------------------------------------
@login_required
def listsub_process(request):
    request.session.set_expiry(3600)
    acc_list = user_details.objects.values('account_name').filter(email_id=request.user.username).distinct()
    proj_list = user_details.objects.values('project_name').filter(email_id=request.user.username).distinct()
    sub_process = ""
    for x in proj_list:
        grp = sub_process_list.objects.filter(project_name=x['project_name'])
        sub_process = list(chain(sub_process, grp))

    data = {}
    data['accounts'] = acc_list
    data['sub_process'] = sub_process

    return render(request, 'e4-sub_process.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- CREATE SUB PROCESS -----------------------------------------------------------------
@login_required
def create_sub_process(request):

    data = {}
    request.session.set_expiry(3600)
    ac_name     = request.POST['account_name']
    proj_name   = request.POST['project_name']
    sp_name     = request.POST['sub_process_name'].strip()

    check = sub_process_list.objects.filter(sub_process_name = sp_name).count()

    if check == 0:
        sub_process = sub_process_list(sub_process_name=sp_name, account_name=ac_name, project_name=proj_name)
        sub_process.save()

        data['message'] = "Sub Process created Successfully"
    else:
        data['message'] = "Sub Process already exists"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE SUB PROCESS -----------------------------------------------------------------
@login_required
def delete_sub_process(request):
    request.session.set_expiry(3600)
    sub_process_id = request.POST['id']
    sb_process = sub_process_list.objects.get(pk=sub_process_id)
    sub_process_list.objects.get(pk=sub_process_id).delete()
    sub_process_details.objects.filter(sub_process_name=sb_process.sub_process_name).delete()
    return render(request, 'e4-sub_process.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ADD USER In SUBPROCESS -------------------------------------------------------------
@login_required
def sub_process_add_user(request):
    request.session.set_expiry(3600)
    sp_name = request.POST['sub_process_name']
    usr_email = json.loads(request.POST['user_email'])

    account_name = sub_process_list.objects.filter(sub_process_name = sp_name)[0].account_name
    project_name = sub_process_list.objects.filter(sub_process_name = sp_name)[0].project_name

    user_list = []

    usr_email_count = len(usr_email)
    for i in range(usr_email_count):
        user = {}
        user['ac_name']     = account_name
        user['proj_name']   = project_name
        user['sp_name']     = sp_name.strip()
        user['usr_email']   = usr_email[i].strip()
        user_list.append(user)

    add_user_subprocess(user_list)

    return render(request, 'e4-sub_process.html')

def add_user_subprocess(user_list):
    for x in user_list:
        sub_process = sub_process_details(account_name      = x['ac_name'],
                                          project_name      = x['proj_name'],
                                          sub_process_name  = x['sp_name'],
                                          user_email        = x['usr_email'])
        sub_process.save()
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE USER FROM SUB PROCESS -------------------------------------------------------
@login_required
def sub_process_del_user(request):
    request.session.set_expiry(3600)

    user_id_list = json.loads(request.POST['id'])

    for sub_process_user_id in user_id_list:
        sub_process_details.objects.get(pk=sub_process_user_id).delete()

    return render(request, 'e4-sub_process.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST DROPDOWN SUB PROCESS ----------------------------------------------------------
@login_required
def load_sub_process(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        account_name = request.POST.get('account_name')
        project_name = request.POST.get('project_name')

        data['sub_process_list'] = []
        sp_list = sub_process_list.objects.filter(account_name=account_name, project_name=project_name)
        for sp in sp_list:
            set = {}
            set['sub_process_name'] = sp.sub_process_name
            set['account_name']     = sp.sub_process_name
            set['project_name']     = sp.sub_process_name
            data['sub_process_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST USER OF SELECTED SUB PROCESS --------------------------------------------------
def load_sub_process_users(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        subprocess_name = request.POST.get('sub_process_name')

        data['sp_users_list'] = []
        sp_users = sub_process_details.objects.filter(sub_process_name=subprocess_name)
        for sp_u in sp_users:
            set = {}
            set['id']               = sp_u.id
            set['account_name']     = sp_u.account_name
            set['project_name']     = sp_u.project_name
            set['sub_process_name'] = sp_u.sub_process_name
            set['user_email']       = sp_u.user_email
            data['sp_users_list'].append(set)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST DROPDOWN OF USERS FROM SELECTED PROJECT ---------------------------------------
def load_project_users(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        project_name = request.POST.get('project_name')

        users_list = user_details.objects.filter(project_name=project_name).values('email_id').distinct()
        data = json.dumps(list(users_list))

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- LIST DROPDOWN OF USERS FROM SELECTED PROJECT ---------------------------------------
def load_project_users_from_sp(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}
    if csrf_token != None:

        data = []
        request.session.set_expiry(3600)
        sp_name = request.POST.get('sub_process')
        project_name = sub_process_list.objects.filter(sub_process_name = sp_name)[0].project_name

        users_list = user_details.objects.filter(project_name=project_name).values('email_id').distinct()

        for usr in users_list:

            role_check = User.objects.filter(username = usr['email_id'])[0].email
            if role_check == "agent" or role_check == "superuser":

                usr_avail_count = sub_process_details.objects.filter(sub_process_name = sp_name, user_email = usr['email_id']).count()
                if usr_avail_count == 0:

                    set = {}
                    set['email_id'] = usr['email_id']
                    data.append(set)
    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ADD USERS In BATCH IN SUBPROCESS BY UPLOADING EXCEL-SHEET --------------------------
def upload_subprocess_users_file(request):
    if request.method == 'POST' and request.FILES['user_file']:
        request.session.set_expiry(3600)
        myfile = request.FILES['user_file']
        fs = FileSystemStorage('media/excel_uploads/sub_process/')
        filename = fs.save((myfile.name).replace(" ", "_"), myfile)
        uploaded_file_url = fs.url("/excel_uploads/sub_process/" + filename)
        create_subprocess_users_from_excel(uploaded_file_url)

    return redirect('/dashboard/sub_process/')

def create_subprocess_users_from_excel(file_path):
    book = xlrd.open_workbook(str(settings.BASE_DIR) + file_path)
    sheet = book.sheet_by_index(0)
    user_list = []
    # loop over each row
    for r in range(1, sheet.nrows):
        ac_name     = (sheet.cell(r, 0).value).strip()
        proj_name   = (sheet.cell(r, 1).value).strip()
        sp_name     = (sheet.cell(r, 2).value).strip()
        user_email  = (sheet.cell(r, 3).value).strip()

        if User.objects.filter(username = user_email).exists():
            user_type = User.objects.filter(username = user_email).values('email')[0]['email']

            if user_type == "agent" or user_type == "superuser":
                if user_details.objects.filter(email_id = user_email, account_name = ac_name, project_name = proj_name).exists():
                    if sub_process_list.objects.filter(sub_process_name = sp_name, account_name = ac_name, project_name = proj_name).exists():
                        # extract each cell
                        user = {}
                        user['ac_name']     = ac_name
                        user['proj_name']   = proj_name
                        user['sp_name']     = sp_name
                        user['usr_email']   = user_email

                        user_list.append(user)

    add_user_subprocess(user_list)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DOWNLOAD BULK-UPLOAD IN SUBPROCESS TEMPLATE ----------------------------------------
@login_required
def download_sp_user_template(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')

    if csrf_token != None:
        request.session.set_expiry(3600)
        path = '/excel_upload_templates/upload_subprocess_users.xlsx'
        file_path = str(settings.MEDIA_ROOT) + path
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Sub Process (Create Sub Process) -----------------------------------------------





# --------------------------------- Add Account ------------------------------------------------------------------------


# --------------------------------- REDIRECT TO ADD ACCOUNT ------------------------------------------------------------
@login_required
def listAccount(request):
    request.session.set_expiry(3600)
    acc_list = account_list.objects.all().order_by('account_name')
    return render(request, 'f1-create-account.html', {'acc_list': acc_list})
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ADD ACCOUNT ------------------------------------------------------------------------
def sup_add_account(request):

    data = {}
    request.session.set_expiry(3600)
    ac_name = request.POST['accountname'].strip()
    check = account_list.objects.filter(account_name=ac_name).count()

    if check == 0:
        acc = account_list(account_name=ac_name)
        acc.save()
        data['message'] = "Account created Successfully"
    else:
        data['message'] = "Account already exists"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE ACCOUNT ---------------------------------------------------------------------
def sup_del_account(request):
    request.session.set_expiry(3600)

    ac_id   = request.POST['id']
    acc     = account_list.objects.filter(pk = ac_id)[0].account_name

    if user_details.objects.filter(account_name = acc).exists():
        user_details.objects.filter(account_name = acc).delete()

    if sub_process_list.objects.filter(account_name = acc).exists():
        sub_process_list.objects.filter(account_name = acc).delete()

    if sub_process_details.objects.filter(account_name = acc).exists():
        sub_process_details.objects.filter(account_name = acc).delete()

    if projects_list.objects.filter(account_name = acc).exists():
        projects_list.objects.filter(account_name = acc).delete()

    account_list.objects.get(pk=ac_id).delete()
    return render(request, 'f1-create-account.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Add Account --------------------------------------------------------------------





# --------------------------------- Add Project ------------------------------------------------------------------------


# --------------------------------- REDIRECT TO ADD PROJECT ------------------------------------------------------------
@login_required
def listProject(request):
    request.session.set_expiry(3600)
    data = {}

    proj_list = projects_list.objects.all().order_by('account_name', 'project_name')
    acc_list = account_list.objects.all()
    data['projects'] = proj_list
    data['accounts'] = acc_list

    return render(request, 'f2-create-project.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ADD PROJECT ------------------------------------------------------------------------
def sup_add_project(request):

    data = {}
    request.session.set_expiry(3600)
    ac_name = request.POST['accountname']
    proj_name = request.POST['projectname'].strip()
    check = projects_list.objects.filter(project_name=proj_name).count()

    if check == 0:
        proj = projects_list(account_name=ac_name, project_name=proj_name)
        proj.save()

        admins_list = User.objects.filter(Q(email = "superuser") | Q(email = "leader")).values('username')
        for u in admins_list:
            user_data = user_details.objects.filter(email_id = u['username'])[0]
            emp_id      = user_data.emp_id
            emp_name    = user_data.full_name
            emp_email   = user_data.email_id
            save_admin_details(ac_name, proj_name, emp_id, emp_name, emp_email)

        user_list = user_details.objects.filter(account_name = ac_name)
        for usr in user_list:
            user_type = User.objects.filter(username=usr.email_id).values('email')[0]['email']
            if user_type == "admin":
                emp_id      = usr.emp_id
                emp_name    = usr.full_name
                emp_email   = usr.email_id
                save_admin_details(ac_name, proj_name, emp_id, emp_name, emp_email)

        data['message'] = "Project created Successfully"
    else:
        data['message'] = "Project already exists"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION------------------------------------------------------------------------


# --------------------------------- DELETE SELECTED PROJECT ------------------------------------------------------------
def sup_del_project(request):
    request.session.set_expiry(3600)

    proj_id = request.POST['id']
    proj = projects_list.objects.filter(pk = proj_id)[0].project_name

    if user_details.objects.filter(project_name = proj).exists():
        user_details.objects.filter(project_name = proj).delete()

    if sub_process_list.objects.filter(project_name = proj).exists():
        sub_process_list.objects.filter(project_name = proj).delete()

    if sub_process_details.objects.filter(project_name = proj).exists():
        sub_process_details.objects.filter(project_name = proj).delete()

    projects_list.objects.get(pk=proj_id).delete()

    return render(request, 'f2-create-project.html')
# --------------------------------- END FUNCTION------------------------------------------------------------------------


# --------------------------------- END Add Project --------------------------------------------------------------------





# --------------------------------- Create Admin -----------------------------------------------------------------------


# --------------------------------- REDIRECT TO CREATE ADMIN -----------------------------------------------------------
@login_required
def listAdmin(request):
    request.session.set_expiry(3600)
    data = {}

    users_info = []
    users_list = User.objects\
                    .filter(Q(email = "superuser") | Q(email = "leader") | Q(email = "admin") | Q(email = "poc") | Q(email = "agent"))\
                    .values('id', 'first_name', 'last_name', 'username', 'email').order_by('first_name')

    for u in users_list:

        emp_id = user_details.objects.filter(email_id = u['username']).values('emp_id')[0]['emp_id']

        set                 = {}
        set['id']           = u['id']
        set['emp_id']       = emp_id
        set['first_name']   = u['first_name']
        set['last_name']    = u['last_name']
        set['email']        = u['username']
        set['user_type']    = u['email']
        users_info.append(set)

    acc_list = account_list.objects.all()

    data['admins'] = users_info
    data['accounts'] = acc_list


    return render(request, 'f3-create-admin.html', data)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- CREATE ADMIN WITH THE ACCOUNT And PROJECT DETAILS PROVIDED -------------------------
def sup_create_admin(request):
    request.session.set_expiry(3600)
    admin_acname = request.POST['account_name']
    admin_projname = request.POST['project_name']
    admin_emp_id = request.POST['emp_id']
    admin_firstname = request.POST['first_name']
    admin_lastname = request.POST['last_name']
    admin_fullname = admin_firstname + " " + admin_lastname
    admin_email = request.POST['email_id']
    admin_type = request.POST['admin_type']
    password = "cognizant"

    regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'

    if account_list.objects.filter(account_name=admin_acname).exists() and \
            projects_list.objects.filter(account_name=admin_acname, project_name=admin_projname).exists() and \
            len(admin_emp_id) > 5 and \
            len(admin_emp_id) < 8 and \
            admin_emp_id.isdigit() and \
            all(x.isalpha() or x.isspace() for x in admin_firstname) and \
            all(x.isalpha() or x.isspace() for x in admin_lastname) and \
            (re.search(regex, admin_email)):

        if admin_type == "superuser" or \
                admin_type == "leader" or \
                admin_type == "admin" or \
                admin_type == "poc" or \
                admin_type == "agent":

            if admin_type == "superuser" or admin_type == "leader":
                proj_list = projects_list.objects.all()
                for p in proj_list:
                    save_admin_details(p.account_name, p.project_name, admin_emp_id, admin_fullname, admin_email)

            elif admin_type == "admin":
                proj_list = projects_list.objects.filter(account_name = admin_acname)
                for p in proj_list:    
                    save_admin_details(p.account_name, p.project_name, admin_emp_id, admin_fullname, admin_email)

            else:
                save_admin_details(admin_acname, admin_projname, admin_emp_id, admin_fullname, admin_email)

            if User.objects.filter(username=admin_email).exists():
                pass
            else:
                u = User.objects.create_user(username=admin_email, password=password,
                                             first_name=admin_firstname, last_name=admin_lastname, email=admin_type)

                u.save()

            sender = request.user.username

            receiver = []
            receiver.append(admin_email)

            send_email_for_user_creation(sender, admin_firstname, receiver, admin_type)

    return render(request, 'f3-create-admin.html')

# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- SAVE ADMIN DETAILS -----------------------------------------------------------------
def save_admin_details(account_name, project_name, emp_id, emp_name, emp_email):
    if user_details.objects.\
            filter(account_name = account_name, project_name = project_name, email_id = emp_email).exists():
        pass
    else:
        admin_details = user_details(account_name = account_name,
                                     project_name = project_name,
                                     emp_id = emp_id,
                                     full_name = emp_name,
                                     email_id = emp_email)
        admin_details.save()
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- DELETE ADMIN FROM THE ACCOUNT And PROJECT DETAILS ASSOCIATED -----------------------
def sup_del_admin_entry(request):
    request.session.set_expiry(3600)

    user_id_list = json.loads(request.POST['id'])

    for user_id in user_id_list:

        user_email = User.objects.filter(pk = user_id).values('username')[0]['username']

        user_details_list = user_details.objects.filter(email_id = user_email)
        for p in user_details_list:
            delete_user_f_proj_sp(p.project_name, user_email)
            user_details.objects.filter(project_name = p.project_name, email_id = user_email).delete()

        User.objects.get(pk=user_id).delete()

    return render(request, 'f3-create-admin.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------

# --------------------------------- Change Update validity -------------------------------------------------------------
@login_required
def change_user_type(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)

        users_list = json.loads(request.POST['users_list'])
        user_type = request.POST.get('user_type')

        users_count = len(users_list)

        for i in range(users_count):
            User.objects.filter(username = users_list[i]).update(email = user_type)


        data['message'] = "User-Type changed successfully"

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------

# --------------------------------- END Create Admin -------------------------------------------------------------------





# --------------------------------- Email Triggers ---------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL ON REGISTERING USER --------------------------------------------------
def cc_list_of_poc_and_admins(sub_process, sender):

    cc_receiver = []
    poc_project_name = sub_process_list.objects.filter(sub_process_name = sub_process).values('project_name')[0]['project_name']
    user_list = user_details.objects.filter(project_name=poc_project_name)
    for usr in user_list:
        user_type = User.objects.filter(username=usr.email_id).values('email')[0]['email']
        if user_type == "poc" or user_type == "admin":
            cc_receiver.append(usr.email_id)
    cc_receiver.append(sender)

    return cc_receiver
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL ON REGISTERING USER --------------------------------------------------
def send_email_for_user_creation(sender, user_firstname, receiver, user_type):
    msg = EmailMessage()
    msg['Subject'] = 'You are successfully registered in Elixir'
    msg['From'] = sender
    msg['To'] = ", ".join(receiver)
    msg['cc'] = sender
    # msg.set_content('')
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/user_registration_successfull.html',
        {"user_f_name": user_firstname,
         "user_type": user_type})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER UPDATE EMAILS TO USERS -----------------------------------------------------
def send_update_mail(sender, receiver, cc_receiver, title, update_source, update_type, message, validity, attachments):
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ", ".join(receiver)
    msg['cc'] = ",".join(cc_receiver)
    msg['Subject'] = title
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_update_details.html',
        {"title": title,
         "update_source": update_source,
         "update_type": update_type,
         "validity": validity}) + message
    msg.attach(MIMEText(html, 'html'))

    if len(attachments) > 0:
        for a_file in attachments:
            attachment = open(str(settings.BASE_DIR) + a_file, 'rb')
            file_name = os.path.basename(a_file)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            part.add_header('Content-Disposition',
                            'attachment',
                            filename=file_name)
            encoders.encode_base64(part)
            msg.attach(part)
    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL WHEN NEW KNOWLEDGE TEST SENT TO ASSOCIATE ----------------------------
def send_pkt_email(sender, reciever, cc_receiver):
    msg = EmailMessage()
    msg['Subject'] = 'New knowledge test for you'
    msg['From'] = sender
    msg['To'] = reciever
    msg['cc'] = ",".join(cc_receiver)
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_knowledge_check.html',
        {"reciever": reciever})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER UPDATE EMAILS TO USERS -----------------------------------------------------
def send_update_kc_mail(sender, receiver, cc_receiver, title, update_source, update_type, message, validity, attachments):
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ", ".join(receiver)
    msg['cc'] = ", ".join(cc_receiver)
    msg['Subject'] = title
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_update_kc_details.html',
        {"title": title,
         "update_source": update_source,
         "update_type": update_type,
         "validity": validity}) + message
    msg.attach(MIMEText(html, 'html'))

    if len(attachments) > 0:
        for a_file in attachments:
            attachment = open(str(settings.BASE_DIR) + a_file, 'rb')
            file_name = os.path.basename(a_file)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            part.add_header('Content-Disposition',
                            'attachment',
                            filename=file_name)
            encoders.encode_base64(part)
            msg.attach(part)
    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL WHEN NEW KNOWLEDGE TEST SENT TO ASSOCIATE ----------------------------
def send_update_pkt_email(sender, reciever, cc_receiver):
    msg = EmailMessage()
    msg['Subject'] = 'New knowledge test for you'
    msg['From'] = sender
    msg['To'] = reciever
    msg['cc'] = ", ".join(cc_receiver)
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_up_knowledge_check.html',
        {"reciever": reciever})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL WHEN ACKNOWLEDGEMENT SUBMITTED BY ASSOCIATE --------------------------
def send_acknowledgement_email(sender, reciever, ack_status, comments):
    if ack_status == "yes":
        ack = "I Acknowledged"
    else:
        ack = "I Need Clarification"

    msg = EmailMessage()
    msg['Subject'] = 'New acknowledgement arrived'
    msg['From'] = sender
    msg['To'] = reciever
    msg['cc'] = sender
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_acknowledgement_received.html',
        {"reciever": reciever,
         "sender": sender,
         "ack_status": ack,
         "comments" : comments})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL WHEN ACKNOWLEDGEMENT SUBMITTED BY ASSOCIATE --------------------------
def send_knowledge_check_acknowledgement_email(sender, reciever):
    msg = EmailMessage()
    msg['Subject'] = 'New Knowledge Test submission'
    msg['From'] = sender
    msg['To'] = reciever
    msg['cc'] = sender
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/new_kc_acknowledgement_received.html',
        {"reciever": reciever,
         "sender": sender})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL FOR PENDING UPDATES TO ACKNOWLEDGE -----------------------------------
def send_reminder(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        update_id   = request.POST.get('update_id')
        sender      = request.user.username
        sub_process = update_details.objects.filter(id = update_id).values('sub_process')[0]['sub_process']
        cc_receiver = cc_list_of_poc_and_admins(sub_process, sender)

        recipient_list = update_recipients.objects.filter(update_id=update_id, ack_status="NA")
        recipients= []
        for x in recipient_list:
            recipients.append(x.user_email)
        
        send_reminder_email(sender, recipients, cc_receiver)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)

def send_reminder_email(sender, recipients, cc_receiver):
    msg = EmailMessage()
    msg['Subject'] = 'ELIXIR REMINDER : Your acknowledgement is pending for process update(s)'
    msg['From'] = sender
    msg['To'] = ", ".join(recipients)
    msg['cc'] = ", ".join(cc_receiver)
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/pending_update_reminder.html',
        {"reciever": "Agent"})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL FOR PENDING Knowledge Test -------------------------------------------
def send_reminder_kc(request):
    csrf_token = request.POST.get('csrfmiddlewaretoken')
    data = {}

    if csrf_token != None:
        request.session.set_expiry(3600)
        conquest_id = request.POST.get('conquest_id')

        sender = request.user.username
        sub_process = conquest_list.objects.filter(id= conquest_id).values('sub_process')[0]['sub_process']
        cc_receiver = cc_list_of_poc_and_admins(sub_process, sender)

        recipient_list = conquest_recipients.objects.filter(conquest_id=conquest_id, attempts=0)
        recipients = []
        for x in recipient_list:
            recipients.append(x.user_email)

        send_reminder_kc_email(sender, recipients, cc_receiver)

    else:
        data['message'] = "** You are not Authorized to view this Data **"

    return JsonResponse(data, safe=False)

def send_reminder_kc_email(sender, recipients, cc_receiver):
    msg = EmailMessage()
    msg['Subject'] = 'ELIXIR REMINDER : Your submission is pending for Knowledge Test(s)'
    msg['From'] = sender
    msg['To'] = ", ".join(recipients)
    msg['cc'] = ", ".join(cc_receiver)
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/pending_kc_reminder.html',
        {"reciever": "Agent"})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- TRIGGER EMAIL WHEN KNOWLEDGE TEST EVALUATED BY POC/SME -----------------------------
def send_kc_eval_check_email(sender, reciever):
    msg = EmailMessage()
    msg['Subject'] = 'You Knowledge Test submission is Evaluated'
    msg['From'] = sender
    msg['To'] = reciever
    msg['cc'] = sender
    html = render_to_string( str(settings.BASE_DIR) + '/templates/email_templates/knowledge_check_evaluation.html',
        {"reciever": reciever,
         "sender": sender})
    msg.add_alternative(html, subtype="html")

    try:
        server = smtplib.SMTP('APACSMTP.CTS.COM', 25)
        server.ehlo()
        server.send_message(msg)
    except Exception as e:
        print("Unable to send the mail.", e)

    return None
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END Email Triggers -----------------------------------------------------------------






# --------------------------------- ERROR HANDLERS ---------------------------------------------------------------------


# --------------------------------- BAD REQUEST ------------------------------------------------------------------------
def error_400_found(request, exception):
    return render(request, '400.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- ACCESS FORBIDDEN -------------------------------------------------------------------
def error_403_found(request, exception):
    return render(request, '403.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- PAGE NOT FOUND ---------------------------------------------------------------------
def error_404_found(request, exception):
    return render(request, '404.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- INTERNAL SERVER ERROR --------------------------------------------------------------
def error_500_found(request):
    return render(request, '500.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- CSRF FAILURE -----------------------------------------------------------------------
def forbidden_errors(request, reason=""):
    return render(request, '403.html')
# --------------------------------- END FUNCTION -----------------------------------------------------------------------


# --------------------------------- END HANDLERS -----------------------------------------------------------------------


#def multithread():
#    print("abc")
#    time.sleep(3.0)
#    multithread()
#
#multithread()
